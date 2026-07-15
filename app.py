#!/usr/bin/env python3
"""
Event Scheduler - A web-based event scheduling and resource allocation system.
Pure Python implementation using built-in http.server and SQLite.
"""

import http.server
import socketserver
import sqlite3
import json
import os
import re
import html
import smtplib
import threading
import time
from urllib.parse import parse_qs, urlparse
from datetime import datetime, date, timedelta
from email.message import EmailMessage
import openpyxl
import tempfile
import shutil
import cgi

DATABASE = os.path.join(os.path.dirname(__file__), 'database', 'scheduler.db')
PORT = 5000
UPLOAD_DIR = os.path.join(os.path.dirname(__file__), 'uploads')

# Hardcoded credentials - Single user
VALID_CREDENTIALS = {
    'admin': 'admin123'
}

# Gmail reminder configuration (set these in the environment before running)
REMINDER_ENABLED = os.environ.get('EVENT_REMINDER_ENABLED', 'true').lower() == 'true'
REMINDER_HOUR = int(os.environ.get('EVENT_REMINDER_HOUR', '18'))
REMINDER_MINUTE = int(os.environ.get('EVENT_REMINDER_MINUTE', '00'))
SMTP_HOST = os.environ.get('GMAIL_SMTP_HOST', 'smtp.gmail.com')
SMTP_PORT = int(os.environ.get('GMAIL_SMTP_PORT', '587'))
SMTP_USERNAME = os.environ.get('GMAIL_SMTP_USERNAME', '')
SMTP_PASSWORD = os.environ.get('GMAIL_SMTP_PASSWORD', '')
SMTP_FROM_EMAIL = os.environ.get('GMAIL_FROM_EMAIL','')
SMTP_FROM_NAME = os.environ.get('GMAIL_FROM_NAME', 'Developer is testing this application')

# Create upload directory
os.makedirs(UPLOAD_DIR, exist_ok=True)

class ThreadedTCPServer(socketserver.ThreadingMixIn, socketserver.TCPServer):
    allow_reuse_address = True
    daemon_threads = True

def get_db():
    os.makedirs(os.path.dirname(DATABASE), exist_ok=True)
    db = sqlite3.connect(DATABASE)
    db.execute('PRAGMA foreign_keys = ON')
    db.row_factory = sqlite3.Row
    return db

def escape_html(value):
    return html.escape(str(value), quote=True) if value is not None else ''

def display_text(value, fallback='-'):
    return escape_html(value if value not in (None, '') else fallback)

def parse_form(body):
    parsed = parse_qs(body, keep_blank_values=True)
    return {key: values[-1] if values else '' for key, values in parsed.items()}

def parse_form_list(body, key):
    return parse_qs(body, keep_blank_values=True).get(key, [])


def build_tomorrow_event_reminders(db):
    tomorrow = (date.today() + timedelta(days=1)).strftime('%Y-%m-%d')
    reminder_date = date.today().strftime('%Y-%m-%d')

    events = db.execute('''
        SELECT e.id, e.event_name, e.event_type, e.description, e.conducted_by,
               e.date, e.participants, e.remarks, ts.start_time, ts.end_time, h.hall_name
        FROM event e
        JOIN time_slot ts ON ts.id = e.timeslot_id
        LEFT JOIN hall h ON h.id = e.hall_id
        WHERE e.date = ?
        ORDER BY ts.period_no, e.event_name
    ''', (tomorrow,)).fetchall()

    if not events:
        return []

    messages = []
    for event in events:
        assigned_faculty = db.execute('''
            SELECT f.faculty_name, f.email
            FROM event_faculty ef
            JOIN faculty f ON f.id = ef.faculty_id
            WHERE ef.event_id = ?
        ''', (event['id'],)).fetchall()

        for faculty in assigned_faculty:
            if not faculty['email']:
                continue

            body_lines = [
                f'Dear {faculty["faculty_name"]},',
                '',
                'This is a reminder that you are assigned to the following event tomorrow:',
                '',
                f'Event: {event["event_name"]}',
                f'Type: {event["event_type"]}',
                f'Date: {event["date"]}',
                f'Time: {event["start_time"]}-{event["end_time"]}',
                f'Hall: {event["hall_name"] or "Not assigned"}',
                f'Conducted by: {event["conducted_by"] or "Not specified"}',
                f'Description: {event["description"] or "No additional description"}',
                f'Participants: {event["participants"] or "Not specified"}',
                f'Remarks: {event["remarks"] or "No remarks"}',
                '',
                'Please make the necessary arrangements accordingly.'
            ]
            message = EmailMessage()
            message['Subject'] = f'Reminder: {event["event_name"]} tomorrow'
            message['From'] = f'{SMTP_FROM_NAME} <{SMTP_FROM_EMAIL}>' if SMTP_FROM_NAME else SMTP_FROM_EMAIL
            message['To'] = faculty['email']
            message.set_content('\n'.join(body_lines))
            messages.append((faculty['email'], message))

    if messages:
        try:
            db.execute('INSERT INTO email_reminder_log (sent_date, summary) VALUES (?, ?)',
                       (reminder_date, f'{len(messages)} reminder(s)'))
            db.commit()
        except sqlite3.IntegrityError:
            db.rollback()

    return messages


def send_tomorrow_event_reminders():
    if not REMINDER_ENABLED:
        return

    if not SMTP_USERNAME or not SMTP_PASSWORD:
        print('Reminder emails skipped: Gmail SMTP credentials are not configured.')
        return

    db = get_db()
    try:
        print('Building reminder messages...')
        messages = build_tomorrow_event_reminders(db)
        print(f'Reminder message count: {len(messages)}')
        if not messages:
            print('No reminder messages were built; skipping send.')
            return False

        print('Connecting to Gmail SMTP...')
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30) as smtp:
            smtp.ehlo()
            smtp.starttls()
            print('Logging into Gmail SMTP...')
            smtp.login(SMTP_USERNAME, SMTP_PASSWORD)
            for _, message in messages:
                print(f'Sending reminder to: {message["To"]}')
                smtp.send_message(message)

        print(f'Reminder emails sent: {len(messages)}')
        return True
    except smtplib.SMTPAuthenticationError as exc:
        print(f'Gmail authentication failed: {exc}')
        return False
    except Exception as exc:
        print(f'Failed to send reminder emails: {exc}')
        return False
    finally:
        db.close()


def run_daily_email_scheduler():
    while True:
        try:
            now = datetime.now()
            target = now.replace(hour=REMINDER_HOUR, minute=REMINDER_MINUTE, second=0, microsecond=0)
            if now > target:
                target += timedelta(days=1)

            sleep_seconds = max(1, int((target - now).total_seconds()))
            if sleep_seconds > 60:
                sleep_seconds = 60
            time.sleep(sleep_seconds)

            current_time = datetime.now()
            if current_time.hour == REMINDER_HOUR and current_time.minute == REMINDER_MINUTE:
                print(f"Reminder scheduler triggered at {current_time.strftime('%Y-%m-%d %H:%M:%S')}")
                send_tomorrow_event_reminders()
                time.sleep(65)
        except Exception as exc:
            print(f'Reminder scheduler error: {exc}')
            time.sleep(60)

def normalize_date_filter(value, fallback=''):
    value = (value or '').strip()
    if not value:
        return fallback
    try:
        datetime.strptime(value, '%Y-%m-%d')
        return value
    except ValueError:
        return fallback

def normalize_int_filter(value):
    value = (value or '').strip()
    if not value:
        return ''
    try:
        return str(int(value))
    except ValueError:
        return ''

def init_db():
    db = get_db()

    db.executescript('''
        CREATE TABLE IF NOT EXISTS branch (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE
        );

        CREATE TABLE IF NOT EXISTS section (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            branch_id INTEGER NOT NULL,
            section_name TEXT NOT NULL,
            FOREIGN KEY (branch_id) REFERENCES branch (id),
            UNIQUE(branch_id, section_name)
        );

        CREATE TABLE IF NOT EXISTS faculty (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            faculty_name TEXT NOT NULL,
            department TEXT NOT NULL,
            email TEXT UNIQUE,
            phone TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS hall (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            hall_name TEXT NOT NULL UNIQUE,
            capacity INTEGER NOT NULL
        );

        CREATE TABLE IF NOT EXISTS time_slot (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            period_no INTEGER NOT NULL UNIQUE,
            start_time TEXT NOT NULL,
            end_time TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS faculty_availability (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            faculty_id INTEGER NOT NULL,
            day_of_week INTEGER NOT NULL,
            timeslot_id INTEGER NOT NULL,
            is_available BOOLEAN DEFAULT 1,
            FOREIGN KEY (faculty_id) REFERENCES faculty (id) ON DELETE CASCADE,
            FOREIGN KEY (timeslot_id) REFERENCES time_slot (id),
            UNIQUE(faculty_id, day_of_week, timeslot_id)
        );

        CREATE TABLE IF NOT EXISTS faculty_busy_slot (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            faculty_id INTEGER NOT NULL,
            timeslot_id INTEGER NOT NULL,
            FOREIGN KEY (faculty_id) REFERENCES faculty (id),
            FOREIGN KEY (timeslot_id) REFERENCES time_slot (id),
            UNIQUE(faculty_id, timeslot_id)
        );

        CREATE TABLE IF NOT EXISTS event (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            event_name TEXT NOT NULL,
            event_type TEXT NOT NULL,
            description TEXT,
            conducted_by TEXT,
            date DATE NOT NULL,
            timeslot_id INTEGER NOT NULL,
            hall_id INTEGER NOT NULL,
            participants INTEGER,
            remarks TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (timeslot_id) REFERENCES time_slot (id),
            FOREIGN KEY (hall_id) REFERENCES hall (id)
        );

        CREATE TABLE IF NOT EXISTS event_branch_section (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            event_id INTEGER NOT NULL,
            branch_id INTEGER NOT NULL,
            section_id INTEGER NOT NULL,
            FOREIGN KEY (event_id) REFERENCES event (id) ON DELETE CASCADE,
            FOREIGN KEY (branch_id) REFERENCES branch (id),
            FOREIGN KEY (section_id) REFERENCES section (id),
            UNIQUE(event_id, branch_id, section_id)
        );

        CREATE TABLE IF NOT EXISTS event_faculty (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            event_id INTEGER NOT NULL,
            faculty_id INTEGER NOT NULL,
            FOREIGN KEY (event_id) REFERENCES event (id) ON DELETE CASCADE,
            FOREIGN KEY (faculty_id) REFERENCES faculty (id),
            UNIQUE(event_id, faculty_id)
        );

        CREATE TABLE IF NOT EXISTS email_reminder_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sent_date TEXT NOT NULL UNIQUE,
            sent_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            summary TEXT
        );
    ''')

    if db.execute('SELECT COUNT(*) FROM branch').fetchone()[0] == 0:
        branches = ['CSE', 'CSE-AI', 'ECE', 'EEE', 'Mechanical', 'Civil', 'IT']
        sections = ['A', 'B', 'C']

        for branch_name in branches:
            db.execute('INSERT INTO branch (name) VALUES (?)', (branch_name,))
            branch_id = db.execute('SELECT last_insert_rowid()').fetchone()[0]

            for section_name in sections:
                db.execute('INSERT INTO section (branch_id, section_name) VALUES (?, ?)',
                           (branch_id, section_name))

        time_slots = [
            (1, '9:00', '10:00'),
            (2, '10:00', '11:00'),
            (3, '11:00', '12:00'),
            (4, '2:00', '3:00'),
            (5, '3:15', '4:15')
        ]

        for period, start, end in time_slots:
            db.execute('INSERT INTO time_slot (period_no, start_time, end_time) VALUES (?, ?, ?)',
                       (period, start, end))

        halls = [
            ('Amriteshwari Hall', 200),
            ('Sudhamani Hall', 150),
            ('E-learning Hall', 100),
            ('Yoga Hall', 50),
            ('Ground', 500),
            ('Classroom', 60),
            ('CP Lab', 50),
            ('CAE Lab', 40),
            ('Simulation Lab', 40),
            ('AI Lab', 40),
            ('ECE Lab', 40),
            ('Hall 1', 200),
            ('Hall 2', 150),
            ('Seminar Hall', 100)
        ]

        for name, capacity in halls:
            db.execute('INSERT INTO hall (hall_name, capacity) VALUES (?, ?)', (name, capacity))

        faculties = [
            ('Dr. Rajesh', 'CSE', 'rajesh@college.edu', '9876543210'),
            ('Dr. Priya', 'CSE', 'priya@college.edu', '9876543211'),
            ('Dr. Naveen', 'ECE', 'naveen@college.edu', '9876543212'),
            ('Dr. Anitha', 'CSE', 'anitha@college.edu', '9876543213'),
            ('Dr. Kumar', 'EEE', 'kumar@college.edu', '9876543214'),
            ('Dr. Sridhar', 'Mechanical', 'sridhar@college.edu', '9876543215'),
            ('Dr. Lakshmi', 'Civil', 'lakshmi@college.edu', '9876543216'),
            ('Dr. Venkat', 'IT', 'venkat@college.edu', '9876543217'),
        ]

        for name, dept, email, phone in faculties:
            db.execute('INSERT INTO faculty (faculty_name, department, email, phone) VALUES (?, ?, ?, ?)',
                       (name, dept, email, phone))

        db.commit()
        print("Database initialized with seed data.")
    else:
        print("Database already initialized.")

    db.close()

# Login template with full CSS styling
LOGIN_TEMPLATE = '''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Event Scheduler - Login</title>
    <link href="https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@400;500;600;700;800&display=swap" rel="stylesheet">
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
    <link href="/static/css/style.css?v=1.1" rel="stylesheet">
</head>
<body class="login-page-body">
    <!-- Animated Background Blobs -->
    <div class="login-blobs">
        <div class="login-blob login-blob-1"></div>
        <div class="login-blob login-blob-2"></div>
        <div class="login-blob login-blob-3"></div>
    </div>

    <main class="container py-5 d-flex align-items-center justify-content-center" style="min-height: 100vh; position: relative; z-index: 2;">
        <div class="login-card">
            <!-- Left Panel: Branding -->
            <div class="login-brand-panel">
                <div class="brand-panel-center">
                    <div class="brand-logo-glow">
                        <svg xmlns="http://www.w3.org/2000/svg" width="38" height="38" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round">
                            <rect x="3" y="4" width="18" height="18" rx="2" ry="2"></rect>
                            <line x1="16" y1="2" x2="16" y2="6"></line>
                            <line x1="8" y1="2" x2="8" y2="6"></line>
                            <line x1="3" y1="10" x2="21" y2="10"></line>
                            <path d="M8 14h.01"></path>
                            <path d="M12 14h.01"></path>
                            <path d="M16 14h.01"></path>
                            <path d="M8 18h.01"></path>
                            <path d="M12 18h.01"></path>
                            <path d="M16 18h.01"></path>
                        </svg>
                    </div>
                    <h1 class="brand-title-large">Event Scheduler</h1>
                    <p class="brand-subtitle-large">Inauguration &amp; Event Planner</p>
                </div>
                
                <div class="brand-panel-footer">
                    <span class="system-badge">v2.0 Stable</span>
                </div>
            </div>

            <!-- Right Panel: Login Form -->
            <div class="login-form-panel">
                <div class="form-header">
                    <h2>Welcome Back</h2>
                    <p class="form-subtitle">Enter your details to manage the portal</p>
                </div>
                
                <form action="/login" id="loginForm" method="POST" onsubmit="handleFormSubmit(event)">
                    <div class="mb-3">
                        <label for="username" class="form-label">Username</label>
                        <div class="input-group login-input-group">
                            <span class="input-group-text">
                                <svg xmlns="http://www.w3.org/2000/svg" width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
                                    <path d="M20 21v-2a4 4 0 0 0-4-4H8a4 4 0 0 0-4 4v2"></path>
                                    <circle cx="12" cy="7" r="4"></circle>
                                </svg>
                            </span>
                            <input type="text" class="form-control" id="username" name="username" placeholder="admin" required autocomplete="username">
                        </div>
                    </div>
                    
                    <div class="mb-4">
                        <label for="password" class="form-label">Password</label>
                        <div class="input-group login-input-group">
                            <span class="input-group-text">
                                <svg xmlns="http://www.w3.org/2000/svg" width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
                                    <rect x="3" y="11" width="18" height="11" rx="2" ry="2"></rect>
                                    <path d="M7 11V7a5 5 0 0 1 10 0v4"></path>
                                </svg>
                            </span>
                            <input type="password" class="form-control" id="password" name="password" placeholder="••••••••" required autocomplete="current-password">
                            <button class="btn toggle-password" type="button" onclick="togglePasswordVisibility()" aria-label="Toggle Password Visibility">
                                <svg xmlns="http://www.w3.org/2000/svg" id="eyeIcon" width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
                                    <path d="M1 12s4-8 11-8 11 8 11 8-4 8-11 8-11-8-11-8z"></path>
                                    <circle cx="12" cy="12" r="3"></circle>
                                </svg>
                            </button>
                        </div>
                    </div>
                    
                    <div class="error-container">
                        {error_message}
                    </div>
                    
                    <button type="submit" class="btn btn-login w-100 py-3 mt-2" id="submitBtn">
                        <span id="btnText">Sign In</span>
                        <span id="btnSpinner" class="spinner-border spinner-border-sm ms-2 d-none" role="status" aria-hidden="true"></span>
                    </button>
                </form>
                
                <div class="demo-credentials-section">
                    <div class="demo-credentials-header">
                        <div class="line"></div>
                        <span class="text">Or Use Credentials</span>
                        <div class="line"></div>
                    </div>
                    <button type="button" class="btn btn-demo-autofill w-100" id="autofillBtn" onclick="autofillDemo()">
                        <span class="flash-icon">⚡</span> Auto-fill Demo Account
                    </button>
                </div>
            </div>
        </div>
    </main>

    <footer class="py-3 mt-auto" style="position: relative; z-index: 2; background: rgba(255, 255, 255, 0.4) !important; backdrop-filter: blur(10px); border-top: 1px solid rgba(0, 0, 0, 0.03);">
        <div class="container text-center text-muted">
            <small>Smart Event Scheduling &amp; Resource Allocation System</small>
        </div>
    </footer>

    <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/js/bootstrap.bundle.min.js"></script>
    <script>
        function autofillDemo() {{
            const userField = document.getElementById('username');
            const passField = document.getElementById('password');
            const autofillBtn = document.getElementById('autofillBtn');
            
            autofillBtn.disabled = true;
            autofillBtn.classList.add('typing');
            
            userField.value = '';
            passField.value = '';
            
            const usernameVal = 'admin';
            const passwordVal = 'admin123';
            
            let i = 0;
            function typeUsername() {{
                if (i < usernameVal.length) {{
                    userField.value += usernameVal.charAt(i);
                    i++;
                    setTimeout(typeUsername, 80);
                }} else {{
                    i = 0;
                    setTimeout(typePassword, 120);
                }}
            }}
            
            function typePassword() {{
                if (i < passwordVal.length) {{
                    passField.value += passwordVal.charAt(i);
                    i++;
                    setTimeout(typePassword, 80);
                }} else {{
                    autofillBtn.disabled = false;
                    autofillBtn.classList.remove('typing');
                    
                    userField.classList.add('autofilled');
                    passField.classList.add('autofilled');
                    setTimeout(() => {{
                        userField.classList.remove('autofilled');
                        passField.classList.remove('autofilled');
                    }}, 800);
                }}
            }}
            
            typeUsername();
        }}

        function togglePasswordVisibility() {{
            const passField = document.getElementById('password');
            const eyeIcon = document.getElementById('eyeIcon');
            
            if (passField.type === 'password') {{
                passField.type = 'text';
                eyeIcon.innerHTML = `
                    <path d="M17.94 17.94A10.07 10.07 0 0 1 12 20c-7 0-11-8-11-8a18.45 18.45 0 0 1 5.06-5.94M9.9 4.24A9.12 9.12 0 0 1 12 4c7 0 11 8 11 8a18.5 18.5 0 0 1-2.16 3.19m-6.72-1.07a3 3 0 1 1-4.24-4.24"></path>
                    <line x1="1" y1="1" x2="23" y2="23"></line>
                `;
            }} else {{
                passField.type = 'password';
                eyeIcon.innerHTML = `
                    <path d="M1 12s4-8 11-8 11 8 11 8-4 8-11 8-11-8-11-8z"></path>
                    <circle cx="12" cy="12" r="3"></circle>
                `;
            }}
        }}

        function handleFormSubmit(e) {{
            const submitBtn = document.getElementById('submitBtn');
            const btnText = document.getElementById('btnText');
            const btnSpinner = document.getElementById('btnSpinner');
            
            submitBtn.disabled = true;
            btnText.textContent = 'Verifying...';
            btnSpinner.classList.remove('d-none');
        }}
    </script>
</body>
</html>'''

TEMPLATES = {
    'base.html': '''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{title}</title>
    <link href="https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@400;500;600;700;800&display=swap" rel="stylesheet">
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet">
    <link href="/static/css/style.css?v=1.1" rel="stylesheet">
</head>
<body>
    <nav class="navbar navbar-expand-lg navbar-dark bg-primary">
        <div class="container">
            <a class="navbar-brand" href="/"><strong>Event Scheduler</strong></a>
            <button class="navbar-toggler" type="button" data-bs-toggle="collapse" data-bs-target="#navbarNav">
                <span class="navbar-toggler-icon"></span>
            </button>
            <div class="collapse navbar-collapse" id="navbarNav">
                <ul class="navbar-nav ms-auto">
                    <li class="nav-item"><a class="nav-link" href="/">Dashboard</a></li>
                    <li class="nav-item"><a class="nav-link" href="/create-event">Create Event</a></li>
                    <li class="nav-item"><a class="nav-link" href="/events">Events</a></li>
                    <li class="nav-item"><a class="nav-link" href="/faculty">Faculty</a></li>
                    <li class="nav-item"><a class="nav-link" href="/halls">Halls</a></li>
                    <li class="nav-item"><a class="nav-link" href="/upload-excel">Import Excel</a></li>
                    <li class="nav-item"><a class="nav-link" href="/logout">Logout</a></li>
                </ul>
            </div>
        </div>
    </nav>
    <main class="container py-5">
        {flash}
        {content}
    </main>
    <footer class="bg-light py-3 mt-auto">
        <div class="container text-center text-muted">
            <small>Smart Event Scheduling &amp; Resource Allocation System</small>
        </div>
    </footer>
    <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/js/bootstrap.bundle.min.js"></script>
</body>
</html>''',

    'dashboard': '''
<div class="dashboard-hero mb-4">
    <div>
        <p class="dashboard-kicker">Induction Portal</p>
        <h2 class="mb-1">Academic Dashboard</h2>
        <p class="text-muted mb-0">Control center for induction schedules, room bookings, and quick actions.</p>
    </div>
    <a href="/create-event" class="btn btn-primary">
        <svg xmlns="http://www.w3.org/2000/svg" width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5" class="me-1">
            <line x1="12" y1="5" x2="12" y2="19"></line>
            <line x1="5" y1="12" x2="19" y2="12"></line>
        </svg>
        Schedule Event
    </a>
</div>

<!-- Metrics Row -->
<div class="metrics-row mb-4">
    <div class="metric-card">
        <div class="metric-icon">
            <svg xmlns="http://www.w3.org/2000/svg" width="24" height="24" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
                <rect x="3" y="4" width="18" height="18" rx="2" ry="2"></rect>
                <line x1="16" y1="2" x2="16" y2="6"></line>
                <line x1="8" y1="2" x2="8" y2="6"></line>
                <line x1="3" y1="10" x2="21" y2="10"></line>
            </svg>
        </div>
        <div class="metric-info">
            <div class="metric-value">{today_count}</div>
            <div class="metric-label">Today's Programs</div>
        </div>
    </div>
    <div class="metric-card">
        <div class="metric-icon">
            <svg xmlns="http://www.w3.org/2000/svg" width="24" height="24" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
                <path d="M17 21v-2a4 4 0 0 0-4-4H5a4 4 0 0 0-4 4v2"></path>
                <circle cx="9" cy="7" r="4"></circle>
                <path d="M23 21v-2a4 4 0 0 0-3-3.87"></path>
                <path d="M16 3.13a4 4 0 0 1 0 7.75"></path>
            </svg>
        </div>
        <div class="metric-info">
            <div class="metric-value">{next7_count}</div>
            <div class="metric-label">Next 7 Days</div>
        </div>
    </div>
    <div class="metric-card">
        <div class="metric-icon">
            <svg xmlns="http://www.w3.org/2000/svg" width="24" height="24" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
                <circle cx="12" cy="12" r="10"></circle>
                <polyline points="12 6 12 12 16 14"></polyline>
            </svg>
        </div>
        <div class="metric-info">
            <div class="metric-value">{total_events}</div>
            <div class="metric-label">Total Programs</div>
        </div>
    </div>
    <div class="metric-card">
        <div class="metric-icon">
            <svg xmlns="http://www.w3.org/2000/svg" width="24" height="24" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
                <path d="M3 9l9-7 9 7v11a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2z"></path>
                <polyline points="9 22 9 12 15 12 15 22"></polyline>
            </svg>
        </div>
        <div class="metric-info">
            <div class="metric-value">{available_halls_today}/{hall_count}</div>
            <div class="metric-label">Halls Free Today</div>
        </div>
    </div>
</div>

<div class="row g-4 mb-4">
    <!-- Left Column: Calendar Widget & Date details -->
    <div class="col-lg-7">
        <div class="card h-100">
            <div class="card-header border-bottom-0 bg-white pt-4 px-4 d-flex justify-content-between align-items-center">
                <h5 class="mb-0 fw-bold">Interactive Calendar</h5>
                <div class="d-flex align-items-center gap-2">
                    <button class="btn btn-sm btn-outline-secondary px-2" id="prevMonth">&laquo;</button>
                    <span id="currentMonth" class="fw-bold px-2 text-dark" style="min-width: 120px; text-align: center;"></span>
                    <button class="btn btn-sm btn-outline-secondary px-2" id="nextMonth">&raquo;</button>
                </div>
            </div>
            <div class="card-body px-4 pb-4">
                <div id="calendar" class="calendar-grid"></div>
                
                <!-- Expanded Day Details directly inside calendar block -->
                <div class="mt-4 pt-4 border-top">
                    <div class="d-flex justify-content-between align-items-center mb-3">
                        <h6 class="fw-bold mb-0 text-dark">Schedule for <span id="selectedDateText" class="text-primary">-</span></h6>
                        <span class="badge bg-secondary" id="selectedDateCount">0 event(s)</span>
                    </div>
                    <div id="selectedDateEvents" class="dashboard-selected-events">
                        <p class="text-muted small mb-0">Select any highlighted date on the calendar above to list its scheduled events.</p>
                    </div>
                </div>
            </div>
        </div>
    </div>

    <!-- Right Column: Today's Timeline & Quick Links -->
    <div class="col-lg-5">
        <div class="card mb-4" style="max-height: 480px; overflow-y: auto;">
            <div class="card-header bg-white pt-4 px-4 border-bottom-0"><h5 class="mb-0 fw-bold">Today's Schedule</h5></div>
            <div class="card-body px-4 pb-4">
                {today_timeline}
            </div>
        </div>
        
        <div class="card">
            <div class="card-header bg-white pt-4 px-4 border-bottom-0"><h5 class="mb-0 fw-bold">Quick Actions</h5></div>
            <div class="card-body px-4 pb-4">
                <p class="text-muted small">Manage portal directories and bulk imports quickly.</p>
                <div class="quick-links-grid">
                    <a href="/create-event" class="quick-link-btn">
                        <svg xmlns="http://www.w3.org/2000/svg" width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5">
                            <rect x="3" y="4" width="18" height="18" rx="2" ry="2"></rect>
                            <line x1="16" y1="2" x2="16" y2="6"></line>
                            <line x1="8" y1="2" x2="8" y2="6"></line>
                        </svg>
                        Add Event
                    </a>
                    <a href="/upload-excel" class="quick-link-btn">
                        <svg xmlns="http://www.w3.org/2000/svg" width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5">
                            <path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"></path>
                            <polyline points="17 8 12 3 7 8"></polyline>
                            <line x1="12" y1="3" x2="12" y2="15"></line>
                        </svg>
                        Upload Excel
                    </a>
                    <a href="/faculty" class="quick-link-btn">
                        <svg xmlns="http://www.w3.org/2000/svg" width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5">
                            <path d="M17 21v-2a4 4 0 0 0-4-4H5a4 4 0 0 0-4 4v2"></path>
                            <circle cx="9" cy="7" r="4"></circle>
                            <path d="M23 21v-2a4 4 0 0 0-3-3.87"></path>
                            <path d="M16 3.13a4 4 0 0 1 0 7.75"></path>
                        </svg>
                        Faculty List
                    </a>
                    <a href="/halls" class="quick-link-btn">
                        <svg xmlns="http://www.w3.org/2000/svg" width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5">
                            <path d="M3 9l9-7 9 7v11a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2z"></path>
                            <polyline points="9 22 9 12 15 12 15 22"></polyline>
                        </svg>
                        Manage Halls
                    </a>
                </div>
            </div>
        </div>
    </div>
</div>

<!-- Bottom Row: Filter & Queue -->
<div class="card mb-4">
    <div class="card-header bg-white pt-4 px-4 border-bottom-0 d-flex justify-content-between align-items-center">
        <h5 class="mb-0 fw-bold">Program Queue & Filter Overview</h5>
    </div>
    <div class="card-body px-4 pb-4">
        <form method="GET" action="/" class="mb-4">
            <div class="row g-3">
                <div class="col-lg-3 col-md-6">
                    <label class="form-label small fw-bold text-muted">From Date</label>
                    <input type="date" name="date_from" class="form-control" value="{date_from}">
                </div>
                <div class="col-lg-3 col-md-6">
                    <label class="form-label small fw-bold text-muted">To Date</label>
                    <input type="date" name="date_to" class="form-control" value="{date_to}">
                </div>
                <div class="col-lg-2 col-md-6">
                    <label class="form-label small fw-bold text-muted">Branch</label>
                    <select name="branch_id" class="form-select">{branch_options}</select>
                </div>
                <div class="col-lg-2 col-md-6">
                    <label class="form-label small fw-bold text-muted">Hall</label>
                    <select name="hall_id" class="form-select">{hall_options}</select>
                </div>
                <div class="col-lg-2 col-md-12">
                    <label class="form-label small fw-bold text-muted">Event Type</label>
                    <select name="event_type" class="form-select">{type_options}</select>
                </div>
            </div>
            <div class="mt-3">
                <label class="form-label small fw-bold text-muted">Keyword Search</label>
                <div class="input-group">
                    <span class="input-group-text bg-white border-end-0 text-muted">
                        <svg xmlns="http://www.w3.org/2000/svg" width="18" height="18" fill="none" stroke="currentColor" stroke-width="2" viewBox="0 0 24 24">
                            <circle cx="11" cy="11" r="8"/>
                            <path d="m21 21-4.3-4.3"/>
                        </svg>
                    </span>
                    <input type="search" name="q" class="form-control border-start-0 ps-0" value="{search_filter}" placeholder="Search program names, keywords, remarks, or faculty coordinators...">
                </div>
            </div>
            <div class="d-flex justify-content-end gap-2 mt-3">
                <a href="/" class="btn btn-outline-secondary">Reset Filters</a>
                <button type="submit" class="btn btn-primary">Apply Filters</button>
            </div>
        </form>
        
        {upcoming_table}
    </div>
</div>

<script>
const eventsByDate = {events_json};
let currentDate = new Date();
let selectedDate = null;

function escapeHtml(value) {{
    return String(value ?? '').replace(/[&<>"']/g, char => ({{
        '&': '&amp;', '<': '&lt;', '>': '&gt;', '"': '&quot;', "'": '&#39;'
    }}[char]));
}}

function renderCalendar() {{
    const year = currentDate.getFullYear();
    const month = currentDate.getMonth();
    const monthNames = ['January', 'February', 'March', 'April', 'May', 'June',
                      'July', 'August', 'September', 'October', 'November', 'December'];
    document.getElementById('currentMonth').textContent = monthNames[month] + ' ' + year;
    const firstDay = new Date(year, month, 1).getDay();
    const daysInMonth = new Date(year, month + 1, 0).getDate();
    const today = new Date(); today.setHours(0, 0, 0, 0);
    let html = '<div class="calendar-header">';
    ['Sun', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat'].forEach(day => {{
        html += `<div class="calendar-day-name">${{day}}</div>`;
    }});
    html += '</div><div class="calendar-body">';
    for (let i = 0; i < firstDay; i++) {{
        html += '<div class="calendar-day empty"></div>';
    }}
    for (let day = 1; day <= daysInMonth; day++) {{
        const dateStr = `${{year}}-${{String(month + 1).padStart(2, '0')}}-${{String(day).padStart(2, '0')}}`;
        const hasEvent = eventsByDate[dateStr] && eventsByDate[dateStr].length > 0;
        const isToday = new Date(year, month, day).getTime() === today.getTime();
        const isSelected = selectedDate === dateStr;
        let classes = 'calendar-day';
        if (hasEvent) classes += ' has-event';
        if (isToday) classes += ' today';
        if (isSelected) classes += ' selected';
        html += `<div class="${{classes}}" data-date="${{dateStr}}">${{day}}${{hasEvent ? '<span class="event-indicator"></span>' : ''}}</div>`;
    }}
    html += '</div>';
    document.getElementById('calendar').innerHTML = html;
    document.querySelectorAll('.calendar-day:not(.empty)').forEach(el => {{
        el.addEventListener('click', function() {{
            document.querySelectorAll('.calendar-day.selected').forEach(d => d.classList.remove('selected'));
            this.classList.add('selected');
            selectedDate = this.dataset.date;
            loadDateEvents(selectedDate);
        }});
    }});
}}

function loadDateEvents(dateStr) {{
    const [year, month, day] = dateStr.split('-');
    const monthNames = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec'];
    document.getElementById('selectedDateText').textContent = `${{parseInt(day)}} ${{monthNames[parseInt(month)-1]}} ${{year}}`;
    fetch(`/api/events-by-date/${{dateStr}}`).then(r => r.json()).then(events => {{
        const container = document.getElementById('selectedDateEvents');
        const countBadge = document.getElementById('selectedDateCount');
        if (countBadge) countBadge.textContent = events.length + ' event(s)';
        if (events.length === 0) {{
            container.innerHTML = '<p class="text-muted small mb-0">No events scheduled on this date.</p>';
        }} else {{
            let html = '<div class="dashboard-selected-events">';
            events.forEach(event => {{
                html += `<a href="/event/${{event.id}}" class="dashboard-selected-event-item">
                    <div class="d-flex w-100 justify-content-between align-items-start mb-1">
                        <h6 class="mb-0 fw-bold text-dark">${{escapeHtml(event.name)}}</h6>
                        <span class="badge bg-secondary small">${{escapeHtml(event.branch_sections)}}</span>
                    </div>
                    <small class="text-muted d-block">${{escapeHtml(event.timeslot)}} | Venue: ${{escapeHtml(event.hall)}} | Faculty: ${{escapeHtml(event.faculty)}}</small>
                </a>`;
            }});
            html += '</div>';
            container.innerHTML = html;
        }}
    }});
}}

document.getElementById('prevMonth').addEventListener('click', () => {{ currentDate.setMonth(currentDate.getMonth() - 1); renderCalendar(); }});
document.getElementById('nextMonth').addEventListener('click', () => {{ currentDate.setMonth(currentDate.getMonth() + 1); renderCalendar(); }});
renderCalendar();
</script>
''',

    'create_event': '''
<div class="create-event-page">
<div class="row mb-4">
    <div class="col-12">
        <h2>Create Event</h2>
        <p class="text-muted">Schedule a new event by selecting date, branch, section, and available time slot.</p>
    </div>
</div>
<div class="row">
    <div class="col-lg-4 mb-4">
        <div class="card">
            <div class="card-header"><h5 class="mb-0">Step 1: Select Date</h5></div>
            <div class="card-body">
                <label for="eventDate" class="form-label">Event Date</label>
                <input type="date" id="eventDate" class="form-control" value="{preselected_date}">
            </div>
        </div>
    </div>
    <div class="col-lg-4 mb-4">
        <div class="card">
            <div class="card-header"><h5 class="mb-0">Step 2: Select Branches & Sections</h5></div>
            <div class="card-body">
                <p class="text-muted small">Select multiple branches and sections (Ctrl+Click for multiple)</p>
                <label for="selectBranches" class="form-label">Branches</label>
                <select id="selectBranches" class="form-select mb-3" multiple style="min-height: 100px;">
                    <option value="">-- Select Branches --</option>
                </select>
                <label for="selectSections" class="form-label">Sections</label>
                <select id="selectSections" class="form-select" multiple style="min-height: 100px;" disabled>
                    <option value="">-- Select Sections --</option>
                </select>
                <div id="selectedClasses" class="mt-2"></div>
            </div>
        </div>
    </div>
    <div class="col-lg-4 mb-4">
        <div class="card">
            <div class="card-header"><h5 class="mb-0">Step 3: Choose Time Slot</h5></div>
            <div class="card-body">
                <div id="slotSelection"><p class="text-muted">Select date, branches, and sections to view available slots</p></div>
            </div>
        </div>
    </div>
</div>
<div class="row" id="eventFormSection" style="display: none;">
    <div class="col-12">
        <div class="card">
            <div class="card-header"><h5 class="mb-0">Step 4: Event Details</h5></div>
            <div class="card-body">
                <form id="eventForm">
                    <input type="hidden" id="selectedSlotId">
                    <div class="row">
                        <div class="col-md-6 mb-3">
                            <label for="eventName" class="form-label">Event Name *</label>
                            <input type="text" id="eventName" class="form-control" required>
                        </div>
                        <div class="col-md-6 mb-3">
                            <label for="eventType" class="form-label">Event Type *</label>
                            <select id="eventType" class="form-select" required>
                                <option value="Workshop">Workshop</option>
                                <option value="Seminar">Seminar</option>
                                <option value="Technical Event">Technical Event</option>
                                <option value="Club Event">Club Event</option>
                                <option value="Guest Lecture">Guest Lecture</option>
                                <option value="Competition">Competition</option>
                                <option value="Other">Other</option>
                            </select>
                        </div>
                    </div>
                    <div class="row">
                        <div class="col-md-6 mb-3">
                            <label for="conductedBy" class="form-label">Conducted By *</label>
                            <input type="text" id="conductedBy" class="form-control" required>
                        </div>
                        <div class="col-md-6 mb-3">
                            <label for="participants" class="form-label">Expected Participants</label>
                            <input type="number" id="participants" class="form-control" min="1">
                        </div>
                    </div>
                    <div class="mb-3">
                        <label for="description" class="form-label">Event Description</label>
                        <textarea id="description" class="form-control" rows="2"></textarea>
                    </div>
                    <div class="row">
                        <div class="col-md-6 mb-3">
                            <label for="selectHall" class="form-label">Venue/Hall *</label>
                            <select id="selectHall" class="form-select" required>
                                <option value="">-- Select Hall --</option>
                            </select>
                        </div>
                        <div class="col-md-6 mb-3">
                            <label for="selectFaculty" class="form-label">Faculty Coordinators (select multiple)</label>
                            <select id="selectFaculty" class="form-select" multiple style="min-height: 100px;">
                                <option value="">-- Select Faculty --</option>
                            </select>
                            <small class="text-muted">Hold Ctrl/Cmd to select multiple faculty members</small>
                        </div>
                    </div>
                    <div class="mb-3">
                        <label for="remarks" class="form-label">Remarks</label>
                        <textarea id="remarks" class="form-control" rows="2"></textarea>
                    </div>
                    <div class="alert alert-danger" id="conflictAlert" style="display: none;"></div>
                    <div class="d-flex gap-2">
                        <button type="submit" class="btn btn-primary" id="submitBtn">Create Event</button>
                        <a href="/" class="btn btn-secondary">Cancel</a>
                    </div>
                </form>
            </div>
        </div>
    </div>
</div>
</div>
<script>
const state = {{
    date: document.getElementById('eventDate').value || null,
    branchIds: [],
    sectionIds: [],
    slotId: null
}};

function escapeHtml(value) {{
    return String(value ?? '').replace(/[&<>"']/g, char => ({{
        '&': '&amp;', '<': '&lt;', '>': '&gt;', '"': '&quot;', "'": '&#39;'
    }}[char]));
}}

async function loadBranches() {{
    const response = await fetch('/api/branches');
    const branches = await response.json();
    const select = document.getElementById('selectBranches');
    branches.forEach(b => select.innerHTML += `<option value="${{b.id}}">${{escapeHtml(b.name)}}</option>`);
}}

document.getElementById('selectBranches').addEventListener('change', async function() {{
    const selectedOptions = Array.from(this.selectedOptions);
    state.branchIds = selectedOptions.map(opt => parseInt(opt.value)).filter(v => v);
    
    document.getElementById('selectSections').innerHTML = '<option value="">-- Select Sections --</option>';
    state.sectionIds = [];
    resetSlots();
    
    if (state.branchIds.length === 0) {{
        document.getElementById('selectSections').disabled = true;
        return;
    }}
    
    const select = document.getElementById('selectSections');
    select.innerHTML = '';
    select.disabled = false;
    
    for (const branchId of state.branchIds) {{
        const response = await fetch(`/api/sections/${{branchId}}`);
        const sections = await response.json();
        sections.forEach(s => {{
            const branchName = document.querySelector(`#selectBranches option[value="${{branchId}}"]`).text;
            select.innerHTML += `<option value="${{branchId}}_${{s.id}}">${{branchName}} - ${{escapeHtml(s.name)}}</option>`;
        }});
    }}
    
    updateSelectedClasses();
}});

document.getElementById('selectSections').addEventListener('change', function() {{
    const selectedOptions = Array.from(this.selectedOptions);
    state.sectionIds = selectedOptions.map(opt => {{
        const [branchId, sectionId] = opt.value.split('_');
        return {{ branchId: parseInt(branchId), sectionId: parseInt(sectionId) }};
    }}).filter(v => v.branchId && v.sectionId);
    
    updateSelectedClasses();
    loadSlots();
}});

function updateSelectedClasses() {{
    const container = document.getElementById('selectedClasses');
    if (state.sectionIds.length === 0) {{
        container.innerHTML = '<span class="text-muted">No classes selected</span>';
        return;
    }}
    let html = '<div class="d-flex flex-wrap gap-1"><strong>Selected:</strong> ';
    state.sectionIds.forEach(({{branchId, sectionId}}) => {{
        const branchOpt = document.querySelector(`#selectBranches option[value="${{branchId}}"]`);
        const branchName = branchOpt ? branchOpt.text : branchId;
        const sectionOpt = document.querySelector(`#selectSections option[value="${{branchId}}_${{sectionId}}"]`);
        const sectionName = sectionOpt ? sectionOpt.text.split(' - ')[1] : sectionId;
        html += `<span class="badge bg-primary">${{escapeHtml(branchName)}}-${{escapeHtml(sectionName)}}</span>`;
    }});
    html += '</div>';
    container.innerHTML = html;
}}

document.getElementById('eventDate').addEventListener('change', function() {{
    state.date = this.value;
    loadSlots();
}});

async function loadSlots() {{
    const slotContainer = document.getElementById('slotSelection');
    const eventForm = document.getElementById('eventFormSection');
    
    if (!state.date || state.sectionIds.length === 0) {{
        slotContainer.innerHTML = '<p class="text-muted">Select date and sections to view available slots</p>';
        eventForm.style.display = 'none';
        return;
    }}
    
    const sectionIds = state.sectionIds.map(s => s.sectionId);
    const branchIds = state.sectionIds.map(s => s.branchId);
    
    const response = await fetch(`/api/free-slots-multiple?date=${{state.date}}&branch_ids=${{branchIds.join(',')}}&section_ids=${{sectionIds.join(',')}}`);
    const slots = await response.json();
    
    if (slots.error) {{
        slotContainer.innerHTML = '<p class="text-danger">Error loading slots</p>';
        eventForm.style.display = 'none';
        return;
    }}
    if (slots.length === 0) {{
        slotContainer.innerHTML = '<p class="text-warning">No available slots for these selections</p>';
        eventForm.style.display = 'none';
        return;
    }}
    let html = '<div class="d-flex flex-wrap gap-2">';
    slots.forEach(slot => html += `<button type="button" class="btn btn-outline-primary slot-btn" data-slot-id="${{slot.id}}">${{escapeHtml(slot.display)}}</button>`);
    html += '</div>';
    slotContainer.innerHTML = html;
    
    document.querySelectorAll('.slot-btn').forEach(btn => {{
        btn.addEventListener('click', function() {{
            document.querySelectorAll('.slot-btn').forEach(b => {{ b.classList.remove('btn-primary'); b.classList.add('btn-outline-primary'); }});
            this.classList.remove('btn-outline-primary');
            this.classList.add('btn-primary');
            state.slotId = this.dataset.slotId;
            document.getElementById('selectedSlotId').value = state.slotId;
            showEventForm();
        }});
    }});
}}

function resetSlots() {{
    document.getElementById('slotSelection').innerHTML = '<p class="text-muted">Select date, branches, and sections to view available slots</p>';
    document.getElementById('eventFormSection').style.display = 'none';
    state.slotId = null;
}}

async function showEventForm() {{
    document.getElementById('eventFormSection').style.display = 'block';
    const [hallsRes, facultyRes] = await Promise.all([
        fetch(`/api/available-halls?date=${{state.date}}&timeslot_id=${{state.slotId}}`),
        fetch(`/api/available-faculty?date=${{state.date}}&timeslot_id=${{state.slotId}}`)
    ]);
    const halls = await hallsRes.json();
    const faculty = await facultyRes.json();
    const hallSelect = document.getElementById('selectHall');
    hallSelect.innerHTML = '<option value="">-- Select Hall --</option>';
    halls.forEach(h => hallSelect.innerHTML += `<option value="${{h.id}}">${{escapeHtml(h.name)}} (Capacity: ${{escapeHtml(h.capacity)}})</option>`);
    const facultySelect = document.getElementById('selectFaculty');
    facultySelect.innerHTML = '';
    faculty.forEach(f => facultySelect.innerHTML += `<option value="${{f.id}}">${{escapeHtml(f.name)}} (${{escapeHtml(f.department)}})</option>`);
    document.getElementById('conflictAlert').style.display = 'none';
}}

document.getElementById('eventForm').addEventListener('submit', async function(e) {{
    e.preventDefault();
    const submitBtn = document.getElementById('submitBtn');
    submitBtn.disabled = true;
    submitBtn.textContent = 'Creating...';
    
    const facultySelect = document.getElementById('selectFaculty');
    const selectedFaculty = Array.from(facultySelect.selectedOptions).map(option => parseInt(option.value));
    
    const data = {{
        event_name: document.getElementById('eventName').value,
        event_type: document.getElementById('eventType').value,
        conducted_by: document.getElementById('conductedBy').value,
        description: document.getElementById('description').value,
        date: state.date,
        sections: state.sectionIds,
        timeslot_id: parseInt(state.slotId),
        hall_id: parseInt(document.getElementById('selectHall').value),
        faculty_ids: selectedFaculty,
        participants: document.getElementById('participants').value || null,
        remarks: document.getElementById('remarks').value
    }};
    
    const conflictCheck = await fetch('/check-conflicts', {{
        method: 'POST',
        headers: {{ 'Content-Type': 'application/json' }},
        body: JSON.stringify(data)
    }});
    const conflictResult = await conflictCheck.json();
    if (conflictResult.conflicts && conflictResult.conflicts.length > 0) {{
        let errorMsg = '<strong>Conflicts detected:</strong><ul>';
        conflictResult.conflicts.forEach(c => {{
            errorMsg += `<li>${{escapeHtml(c.message)}}</li>`;
            if (c.alternatives && c.alternatives.length > 0) {{
                errorMsg += `<li>Available alternatives: ${{c.alternatives.map(a => escapeHtml(a.name)).join(', ')}}</li>`;
            }}
        }});
        errorMsg += '</ul>';
        document.getElementById('conflictAlert').innerHTML = errorMsg;
        document.getElementById('conflictAlert').style.display = 'block';
        submitBtn.disabled = false;
        submitBtn.textContent = 'Create Event';
        return;
    }}
    try {{
        const response = await fetch('/create-event', {{
            method: 'POST',
            headers: {{ 'Content-Type': 'application/json' }},
            body: JSON.stringify(data)
        }});
        const result = await response.json();
        if (result.success) window.location.href = '/events';
        else {{
            if (result.conflicts && result.conflicts.length > 0) {{
                document.getElementById('conflictAlert').innerHTML = '<strong>Conflicts detected:</strong><ul>' + result.conflicts.map(c => `<li>${{escapeHtml(c.message)}}</li>`).join('') + '</ul>';
                document.getElementById('conflictAlert').style.display = 'block';
            }} else {{
                alert('Error: ' + (result.message || 'Failed to create event'));
            }}
            submitBtn.disabled = false;
            submitBtn.textContent = 'Create Event';
        }}
    }} catch (error) {{
        alert('Error creating event. Please try again.');
        submitBtn.disabled = false;
        submitBtn.textContent = 'Create Event';
    }}
}});

loadBranches();
</script>
''',

    'upload_excel': '''
<div class="row mb-4">
    <div class="col-12">
        <h2>Import Events from Excel</h2>
        <p class="text-muted">Upload an Excel file containing the schedule to automatically create events.</p>
    </div>
</div>
<div class="row justify-content-center">
    <div class="col-lg-8">
        <div class="card">
            <div class="card-body">
                <form action="/upload-excel" method="POST" enctype="multipart/form-data">
                    <div class="mb-3">
                        <label for="excelFile" class="form-label">Select Excel File</label>
                        <input type="file" class="form-control" id="excelFile" name="excelFile" accept=".xlsx,.xls" required>
                        <div class="form-text">Upload the schedule Excel file. It should contain sheets for each date with the schedule.</div>
                    </div>
                    <div class="mb-3 form-check">
                        <input type="checkbox" class="form-check-input" id="clearExisting" name="clearExisting">
                        <label class="form-check-label" for="clearExisting">Clear existing events before importing</label>
                    </div>
                    <div class="d-flex gap-2">
                        <button type="submit" class="btn btn-primary">Upload and Import</button>
                        <a href="/" class="btn btn-secondary">Cancel</a>
                    </div>
                </form>
            </div>
        </div>
        <div class="card mt-4">
            <div class="card-header">
                <h5 class="mb-0">Instructions</h5>
            </div>
            <div class="card-body">
                <ol class="mb-0">
                    <li>The Excel file should have sheets named with dates (e.g., "16.7", "17.7", etc.)</li>
                    <li>Each sheet should have the schedule in a grid format with columns for each section</li>
                    <li>The first row should contain the date</li>
                    <li>The second row should contain section headers (e.g., "CSE A", "CSE B", etc.)</li>
                    <li>Each time slot row should have: Event name, Venue, Faculty Assigned</li>
                    <li>Events will be created with the event name, date, time slot, and branch/section</li>
                    <li>Faculty and venue details will be stored in remarks for reference</li>
                </ol>
            </div>
        </div>
    </div>
</div>
''',

    'event_list': '''
<div class="events-list-page">
<div class="page-hero">
    <div class="page-title-block">
        <p class="page-kicker">Events</p>
        <h2 class="mb-1">Scheduled Events</h2>
        <p class="text-muted mb-0">Search, filter, and review programs without losing context.</p>
    </div>
    <div class="page-toolbar">
        <a href="/create-event" class="btn btn-primary">
            <svg xmlns="http://www.w3.org/2000/svg" width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5" class="me-1">
                <line x1="12" y1="5" x2="12" y2="19"></line>
                <line x1="5" y1="12" x2="19" y2="12"></line>
            </svg>
            Create Event
        </a>
    </div>
</div>
<div class="card mb-4">
    <div class="card-body">
        <form method="GET" action="/events" class="list-toolbar">
            <div class="row g-3">
                <div class="col-lg-5 col-md-12">
                    <label class="form-label">Search</label>
                    <div class="input-group">
                        <span class="input-group-text bg-white border-end-0 text-muted">
                            <svg xmlns="http://www.w3.org/2000/svg" width="18" height="18" fill="none" stroke="currentColor" stroke-width="2" viewBox="0 0 24 24">
                                <circle cx="11" cy="11" r="8"/>
                                <path d="m21 21-4.3-4.3"/>
                            </svg>
                        </span>
                        <input type="search" name="q" class="form-control border-start-0 ps-0" value="{search_filter}" placeholder="Search event name, host, remarks, faculty...">
                    </div>
                </div>
                <div class="col-lg-7 col-md-12">
                    <div class="row g-2">
                        <div class="col-sm-2 col-6">
                            <label class="form-label">From</label>
                            <input type="date" name="date_from" class="form-control" value="{date_from}">
                        </div>
                        <div class="col-sm-2 col-6">
                            <label class="form-label">To</label>
                            <input type="date" name="date_to" class="form-control" value="{date_to}">
                        </div>
                        <div class="col-sm-3 col-6">
                            <label class="form-label">Event Type</label>
                            <select name="event_type" class="form-select">{type_options}</select>
                        </div>
                        <div class="col-sm-3 col-6">
                            <label class="form-label">Venue</label>
                            <select name="hall_id" class="form-select">{hall_options}</select>
                        </div>
                        <div class="col-sm-2 col-6">
                            <label class="form-label">Section</label>
                            <select name="section_id" class="form-select">{section_options}</select>
                        </div>
                    </div>
                </div>
            </div>
            <div class="d-flex justify-content-end gap-2 mt-3">
                <a href="/events" class="btn btn-outline-secondary">Reset Filters</a>
                <button type="submit" class="btn btn-primary">Apply Filters</button>
            </div>
        </form>
    </div>
</div>
<div class="card">
    <div class="card-body p-0">
        {events_table}
    </div>
</div>
</div>

<script>
document.addEventListener('DOMContentLoaded', function() {{
    const searchInput = document.querySelector('input[name="q"]');
    const dateFromInput = document.querySelector('input[name="date_from"]');
    const dateToInput = document.querySelector('input[name="date_to"]');
    const typeSelect = document.querySelector('select[name="event_type"]');
    const hallSelect = document.querySelector('select[name="hall_id"]');
    const sectionSelect = document.querySelector('select[name="section_id"]');
    const table = document.getElementById('eventsTable');
    if (!table) return;
    
    const rows = table.querySelectorAll('tbody tr:not(.empty-row)');
    const resultCount = document.getElementById('resultCount');
    
    // Create an empty state row inside the tbody
    const tbody = table.querySelector('tbody');
    const emptyRow = document.createElement('tr');
    emptyRow.className = 'empty-row';
    emptyRow.style.display = 'none';
    emptyRow.innerHTML = '<td colspan="8" class="text-center py-5 text-muted">No events match the current filters.</td>';
    tbody.appendChild(emptyRow);
 
    function filterEvents() {{
        const query = searchInput.value.toLowerCase().trim();
        const dateFrom = dateFromInput.value;
        const dateTo = dateToInput.value;
        const type = typeSelect.value;
        const hallId = hallSelect.value;
        const sectionId = sectionSelect ? sectionSelect.value : '';
        let visibleCount = 0;
 
        rows.forEach(row => {{
            const rowSearch = row.dataset.search || '';
            const rowDate = row.dataset.date || '';
            const rowType = row.dataset.type || '';
            const rowHallId = row.dataset.hallId || '';
            const rowSectionIds = (row.dataset.sectionIds || '').split(',');
 
            const matchesSearch = !query || rowSearch.includes(query);
            const matchesDateFrom = !dateFrom || rowDate >= dateFrom;
            const matchesDateTo = !dateTo || rowDate <= dateTo;
            const matchesType = !type || rowType === type;
            const matchesHall = !hallId || rowHallId === hallId;
            const matchesSection = !sectionId || rowSectionIds.includes(sectionId);
 
            if (matchesSearch && matchesDateFrom && matchesDateTo && matchesType && matchesHall && matchesSection) {{
                row.style.display = '';
                visibleCount++;
            }} else {{
                row.style.display = 'none';
            }}
        }});
 
        emptyRow.style.display = (visibleCount === 0) ? '' : 'none';
        if (resultCount) {{
            resultCount.textContent = visibleCount + ' event(s) shown';
        }}
    }}
 
    [searchInput, dateFromInput, dateToInput].forEach(el => {{
        if (el) el.addEventListener('input', filterEvents);
    }});
    [typeSelect, hallSelect, sectionSelect].forEach(el => {{
        if (el) el.addEventListener('change', filterEvents);
    }});
}});
</script>
''',

    'event_detail': '''
<div class="row mb-4">
    <div class="col-12">
        <nav aria-label="breadcrumb">
            <ol class="breadcrumb">
                <li class="breadcrumb-item"><a href="/events">Events</a></li>
                <li class="breadcrumb-item active">{event_name}</li>
            </ol>
        </nav>
    </div>
</div>
<div class="row">
    <div class="col-lg-8">
        <div class="card mb-4">
            <div class="card-header d-flex justify-content-between align-items-center">
                <h5 class="mb-0">Event Details</h5>
                <div>
                    <a href="/event/{event_id}/edit" class="btn btn-sm btn-outline-primary">Edit</a>
                    <form action="/event/{event_id}/delete" method="POST" style="display: inline;" onsubmit="return confirm('Are you sure?')">
                        <button type="submit" class="btn btn-sm btn-outline-danger">Delete</button>
                    </form>
                </div>
            </div>
            <div class="card-body">
                <div class="row mb-3">
                    <div class="col-md-6">
                        <label class="form-label text-muted">Event Name</label>
                        <p class="fw-bold">{event_name}</p>
                    </div>
                    <div class="col-md-6">
                        <label class="form-label text-muted">Event Type</label>
                        <p><span class="badge bg-secondary">{event_type}</span></p>
                    </div>
                </div>
                <div class="row mb-3">
                    <div class="col-md-6">
                        <label class="form-label text-muted">Conducted By</label>
                        <p>{conducted_by}</p>
                    </div>
                    <div class="col-md-6">
                        <label class="form-label text-muted">Expected Participants</label>
                        <p>{participants}</p>
                    </div>
                </div>
                <div class="mb-3">
                    <label class="form-label text-muted">Description</label>
                    <p>{description}</p>
                </div>
                <div class="mb-3">
                    <label class="form-label text-muted">Remarks</label>
                    <p>{remarks}</p>
                </div>
                <div class="mb-3">
                    <label class="form-label text-muted">Branches & Sections</label>
                    <p>{branches_sections}</p>
                </div>
            </div>
        </div>
    </div>
    <div class="col-lg-4">
        <div class="card mb-4">
            <div class="card-header"><h5 class="mb-0">Schedule</h5></div>
            <div class="card-body">
                <div class="mb-3">
                    <label class="form-label text-muted">Date</label>
                    <p class="fw-bold">{event_date}</p>
                </div>
                <div class="mb-3">
                    <label class="form-label text-muted">Time Slot</label>
                    <p>{timeslot}</p>
                </div>
            </div>
        </div>
        <div class="card">
            <div class="card-header"><h5 class="mb-0">Resources</h5></div>
            <div class="card-body">
                <div class="mb-3">
                    <label class="form-label text-muted">Venue</label>
                    <p class="fw-bold">{hall}</p>
                </div>
                <div class="mb-3">
                    <label class="form-label text-muted">Faculty Coordinators</label>
                    <p>{faculty}</p>
                </div>
            </div>
        </div>
    </div>
</div>
<div class="row">
    <div class="col-12">
        <a href="/events" class="btn btn-secondary">&laquo; Back to Events</a>
    </div>
</div>
''',

    'edit_event': '''
<div class="edit-event-page">
<div class="row mb-4">
    <div class="col-12">
        <h2>Edit Event</h2>
        <p class="text-muted">Modify event "{event_name}"</p>
    </div>
</div>
<div class="row">
    <div class="col-lg-8">
        <div class="card">
            <div class="card-body">
                <form id="editEventForm">
                    <input type="hidden" id="eventId" value="{event_id}">
                    <input type="hidden" id="currentSections" value='{current_sections_json}'>
                    <div class="row">
                        <div class="col-md-6 mb-3">
                            <label for="eventDate" class="form-label">Event Date *</label>
                            <input type="date" id="eventDate" class="form-control" value="{event_date}" required>
                        </div>
                        <div class="col-md-6 mb-3">
                            <label for="slotSelect" class="form-label">Time Slot *</label>
                            <select id="slotSelect" class="form-select" required>
                                {slot_options}
                            </select>
                        </div>
                    </div>
                    <div class="row">
                        <div class="col-md-12 mb-3">
                            <label class="form-label">Branches & Sections *</label>
                            <div class="row">
                                <div class="col-md-6">
                                    <label for="editBranches" class="form-label small">Branches</label>
                                    <select id="editBranches" class="form-select" multiple style="min-height: 100px;">
                                        <option value="">-- Select Branches --</option>
                                    </select>
                                </div>
                                <div class="col-md-6">
                                    <label for="editSections" class="form-label small">Sections</label>
                                    <select id="editSections" class="form-select" multiple style="min-height: 100px;" disabled>
                                        <option value="">-- Select Sections --</option>
                                    </select>
                                </div>
                            </div>
                            <div id="editSelectedClasses" class="mt-2"></div>
                        </div>
                    </div>
                    <hr>
                    <div class="row">
                        <div class="col-md-6 mb-3">
                            <label for="eventName" class="form-label">Event Name *</label>
                            <input type="text" id="eventName" class="form-control" value="{event_name}" required>
                        </div>
                        <div class="col-md-6 mb-3">
                            <label for="eventType" class="form-label">Event Type *</label>
                            <select id="eventType" class="form-select" required>
                                {type_options}
                            </select>
                        </div>
                    </div>
                    <div class="row">
                        <div class="col-md-6 mb-3">
                            <label for="conductedBy" class="form-label">Conducted By *</label>
                            <input type="text" id="conductedBy" class="form-control" value="{conducted_by}" required>
                        </div>
                        <div class="col-md-6 mb-3">
                            <label for="participants" class="form-label">Expected Participants</label>
                            <input type="number" id="participants" class="form-control" value="{participants}" min="1">
                        </div>
                    </div>
                    <div class="mb-3">
                        <label for="description" class="form-label">Event Description</label>
                        <textarea id="description" class="form-control" rows="2">{description}</textarea>
                    </div>
                    <div class="row">
                        <div class="col-md-6 mb-3">
                            <label for="hallSelect" class="form-label">Venue/Hall *</label>
                            <select id="hallSelect" class="form-select" required>
                                {hall_options}
                            </select>
                        </div>
                        <div class="col-md-6 mb-3">
                            <label for="facultySelect" class="form-label">Faculty Coordinators (select multiple)</label>
                            <select id="facultySelect" class="form-select" multiple style="min-height: 100px;">
                                {faculty_options}
                            </select>
                            <small class="text-muted">Hold Ctrl/Cmd to select multiple faculty members</small>
                        </div>
                    </div>
                    <div class="mb-3">
                        <label for="remarks" class="form-label">Remarks</label>
                        <textarea id="remarks" class="form-control" rows="2">{remarks}</textarea>
                    </div>
                    <div class="alert alert-danger" id="conflictAlert" style="display: none;"></div>
                    <div class="d-flex gap-2">
                        <button type="submit" class="btn btn-primary" id="submitBtn">Save Changes</button>
                        <a href="/event/{event_id}" class="btn btn-secondary">Cancel</a>
                    </div>
                </form>
            </div>
        </div>
    </div>
</div>
</div>
<script>
const currentSections = JSON.parse(document.getElementById('currentSections').value || '[]');
const editState = {{
    sectionIds: currentSections,
    slotId: parseInt(document.getElementById('slotSelect').value)
}};

function escapeHtml(value) {{
    return String(value ?? '').replace(/[&<>"']/g, char => ({{
        '&': '&amp;', '<': '&lt;', '>': '&gt;', '"': '&quot;', "'": '&#39;'
    }}[char]));
}}

async function loadEditBranches() {{
    const response = await fetch('/api/branches');
    const branches = await response.json();
    const select = document.getElementById('editBranches');
    const selectedBranchIds = [...new Set(currentSections.map(s => s.branchId))];
    branches.forEach(b => {{
        const selected = selectedBranchIds.includes(b.id) ? 'selected' : '';
        select.innerHTML += `<option value="${{b.id}}" ${{selected}}>${{escapeHtml(b.name)}}</option>`;
    }});
    loadEditSections();
}}

async function loadEditSections() {{
    const branchSelect = document.getElementById('editBranches');
    const selectedOptions = Array.from(branchSelect.selectedOptions);
    const branchIds = selectedOptions.map(opt => parseInt(opt.value)).filter(v => v);
    
    const sectionSelect = document.getElementById('editSections');
    sectionSelect.innerHTML = '';
    
    if (branchIds.length === 0) {{
        sectionSelect.disabled = true;
        sectionSelect.innerHTML = '<option value="">-- Select Sections --</option>';
        updateEditSelectedClasses();
        return;
    }}
    
    sectionSelect.disabled = false;
    
    const allSections = [];
    for (const branchId of branchIds) {{
        const response = await fetch(`/api/sections/${{branchId}}`);
        const sections = await response.json();
        const branchName = document.querySelector(`#editBranches option[value="${{branchId}}"]`).text;
        sections.forEach(s => {{
            allSections.push({{
                value: `${{branchId}}_${{s.id}}`,
                text: `${{branchName}} - ${{s.name}}`,
                branchId: branchId,
                sectionId: s.id
            }});
        }});
    }}
    
    allSections.forEach(s => {{
        const selected = currentSections.some(cs => cs.branchId === s.branchId && cs.sectionId === s.sectionId) ? 'selected' : '';
        sectionSelect.innerHTML += `<option value="${{s.value}}" ${{selected}}>${{escapeHtml(s.text)}}</option>`;
    }});
    
    updateEditSelectedClasses();
}}

document.getElementById('editBranches').addEventListener('change', function() {{
    loadEditSections();
}});

document.getElementById('editSections').addEventListener('change', function() {{
    const selectedOptions = Array.from(this.selectedOptions);
    editState.sectionIds = selectedOptions.map(opt => {{
        const [branchId, sectionId] = opt.value.split('_');
        return {{ branchId: parseInt(branchId), sectionId: parseInt(sectionId) }};
    }}).filter(v => v.branchId && v.sectionId);
    updateEditSelectedClasses();
}});

function updateEditSelectedClasses() {{
    const container = document.getElementById('editSelectedClasses');
    if (editState.sectionIds.length === 0) {{
        container.innerHTML = '<span class="text-muted">No classes selected</span>';
        return;
    }}
    let html = '<div class="d-flex flex-wrap gap-1"><strong>Selected:</strong> ';
    editState.sectionIds.forEach(({{branchId, sectionId}}) => {{
        const branchOpt = document.querySelector(`#editBranches option[value="${{branchId}}"]`);
        const branchName = branchOpt ? branchOpt.text : branchId;
        const sectionOpt = document.querySelector(`#editSections option[value="${{branchId}}_${{sectionId}}"]`);
        const sectionName = sectionOpt ? sectionOpt.text.split(' - ')[1] : sectionId;
        html += `<span class="badge bg-primary">${{escapeHtml(branchName)}}-${{escapeHtml(sectionName)}}</span>`;
    }});
    html += '</div>';
    container.innerHTML = html;
}}

document.getElementById('editEventForm').addEventListener('submit', async function(e) {{
    e.preventDefault();
    const submitBtn = document.getElementById('submitBtn');
    submitBtn.disabled = true;
    submitBtn.textContent = 'Saving...';
    
    const facultySelect = document.getElementById('facultySelect');
    const selectedFaculty = Array.from(facultySelect.selectedOptions).map(option => parseInt(option.value));
    
    const data = {{
        event_name: document.getElementById('eventName').value,
        event_type: document.getElementById('eventType').value,
        conducted_by: document.getElementById('conductedBy').value,
        description: document.getElementById('description').value,
        date: document.getElementById('eventDate').value,
        sections: editState.sectionIds,
        timeslot_id: parseInt(document.getElementById('slotSelect').value),
        hall_id: parseInt(document.getElementById('hallSelect').value),
        faculty_ids: selectedFaculty,
        participants: document.getElementById('participants').value || null,
        remarks: document.getElementById('remarks').value
    }};
    document.getElementById('conflictAlert').style.display = 'none';
    try {{
        const response = await fetch('/event/{event_id}/edit', {{
            method: 'PUT',
            headers: {{ 'Content-Type': 'application/json' }},
            body: JSON.stringify(data)
        }});
        const result = await response.json();
        if (result.success) window.location.href = '/event/{event_id}';
        else {{
            if (result.conflicts && result.conflicts.length > 0) {{
                document.getElementById('conflictAlert').innerHTML = '<strong>Conflicts detected:</strong><ul>' + result.conflicts.map(c => `<li>${{escapeHtml(c.message)}}</li>`).join('') + '</ul>';
                document.getElementById('conflictAlert').style.display = 'block';
            }} else {{
                alert('Error: ' + (result.message || 'Failed to update event'));
            }}
            submitBtn.disabled = false;
            submitBtn.textContent = 'Save Changes';
        }}
    }} catch (error) {{
        alert('Error updating event.');
        submitBtn.disabled = false;
        submitBtn.textContent = 'Save Changes';
    }}
}});

loadEditBranches();
</script>
''',

    'faculty_list': '''
<div class="page-hero">
    <div class="page-title-block">
        <p class="page-kicker">Faculty</p>
        <h2 class="mb-1">Faculty Directory</h2>
        <p class="text-muted mb-0">Search faculty by name, department, or contact details and keep responsibilities organized.</p>
    </div>
    <div class="page-toolbar">
        <a href="/faculty/add" class="btn btn-primary">
            <svg xmlns="http://www.w3.org/2000/svg" width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5" class="me-1">
                <line x1="12" y1="5" x2="12" y2="19"></line>
                <line x1="5" y1="12" x2="19" y2="12"></line>
            </svg>
            Add Faculty
        </a>
    </div>
</div>
<div class="card mb-4">
    <div class="card-body">
        <form method="GET" action="/faculty" class="list-toolbar">
            <div class="row g-3">
                <div class="col-md-7 col-12">
                    <label class="form-label">Search</label>
                    <div class="input-group">
                        <span class="input-group-text bg-white border-end-0 text-muted">
                            <svg xmlns="http://www.w3.org/2000/svg" width="18" height="18" fill="none" stroke="currentColor" stroke-width="2" viewBox="0 0 24 24">
                                <circle cx="11" cy="11" r="8"/>
                                <path d="m21 21-4.3-4.3"/>
                            </svg>
                        </span>
                        <input type="search" name="q" class="form-control border-start-0 ps-0" value="{search_filter}" placeholder="Search name, email, department, phone...">
                    </div>
                </div>
                <div class="col-md-5 col-12">
                    <label class="form-label">Department</label>
                    <select name="department" class="form-select">{dept_options}</select>
                </div>
            </div>
            <div class="d-flex justify-content-end gap-2 mt-3">
                <a href="/faculty" class="btn btn-outline-secondary">Reset Filters</a>
                <button type="submit" class="btn btn-primary">Apply Filters</button>
            </div>
        </form>
    </div>
</div>
<div class="d-flex justify-content-between align-items-center mb-3">
    <span class="result-pill" id="facultyResultCount">{len_faculties} faculty member(s) shown</span>
</div>
<div class="row g-4">{faculty_cards}</div>

<script>
document.addEventListener('DOMContentLoaded', function() {{
    const searchInput = document.querySelector('input[name="q"]');
    const deptSelect = document.querySelector('select[name="department"]');
    const cards = document.querySelectorAll('.faculty-card-container');
    const resultCount = document.getElementById('facultyResultCount');
    
    // Create empty state element
    const rowContainer = document.querySelector('.row.g-4');
    const emptyCol = document.createElement('div');
    emptyCol.className = 'col-12 empty-state-col';
    emptyCol.style.display = 'none';
    emptyCol.innerHTML = '<div class="empty-state"><p class="text-muted mb-0">No faculty members match the current filters.</p></div>';
    rowContainer.appendChild(emptyCol);

    function filterFaculty() {{
        const query = searchInput.value.toLowerCase().trim();
        const dept = deptSelect.value;
        let visibleCount = 0;

        cards.forEach(card => {{
            const rowSearch = card.dataset.search || '';
            const rowDept = card.dataset.department || '';

            const matchesSearch = !query || rowSearch.includes(query);
            const matchesDept = !dept || rowDept === dept;

            if (matchesSearch && matchesDept) {{
                card.style.display = '';
                visibleCount++;
            }} else {{
                card.style.display = 'none';
            }}
        }});

        emptyCol.style.display = (visibleCount === 0) ? 'block' : 'none';
        if (resultCount) {{
            resultCount.textContent = `${{visibleCount}} faculty member(s) shown`;
        }}
    }}

    if (searchInput) searchInput.addEventListener('input', filterFaculty);
    if (deptSelect) deptSelect.addEventListener('change', filterFaculty);
}});
</script>
''',

    'faculty_form': '''
<div class="row mb-4">
    <div class="col-12"><h2>{title}</h2></div>
</div>
<div class="row justify-content-center">
    <div class="col-lg-8">
        <div class="card">
            <div class="card-body">
                <form method="POST">
                    <div class="mb-3">
                        <label class="form-label">Faculty Name *</label>
                        <input type="text" name="faculty_name" class="form-control" required value="{faculty_name}">
                    </div>
                    <div class="mb-3">
                        <label class="form-label">Department *</label>
                        <select name="department" class="form-select" required>
                            {dept_options}
                        </select>
                    </div>
                    <div class="mb-3">
                        <label class="form-label">Email</label>
                        <input type="email" name="email" class="form-control" value="{email}">
                    </div>
                    <div class="mb-3">
                        <label class="form-label">Phone</label>
                        <input type="text" name="phone" class="form-control" value="{phone}">
                    </div>
                    <div class="d-flex gap-2">
                        <button type="submit" class="btn btn-primary">Save</button>
                        <a href="/faculty" class="btn btn-secondary">Cancel</a>
                    </div>
                </form>
            </div>
        </div>
    </div>
</div>
''',

    'faculty_availability': '''
<div class="row mb-4">
    <div class="col-12">
        <h2>Manage Availability: {faculty_name}</h2>
        <p class="text-muted">Set which periods each faculty member is available on each day of the week</p>
        <p class="text-muted small"><strong>Note:</strong> Check the box to mark a faculty member as <strong>UNAVAILABLE</strong> for that time slot.</p>
    </div>
</div>
<div class="row">
    <div class="col-12">
        <div class="card">
            <div class="card-body">
                <form method="POST">
                    <div class="table-responsive">
                        <div class="availability-grid">
                            <div class="availability-header">Period</div>
                            <div class="availability-header">Monday</div>
                            <div class="availability-header">Tuesday</div>
                            <div class="availability-header">Wednesday</div>
                            <div class="availability-header">Thursday</div>
                            <div class="availability-header">Friday</div>
                            <div class="availability-header">Saturday</div>
                            <div class="availability-header">Sunday</div>
                            
                            {availability_rows}
                        </div>
                    </div>
                    <div class="mt-4 d-flex gap-2">
                        <button type="submit" class="btn btn-primary">Save Availability</button>
                        <a href="/faculty" class="btn btn-secondary">Cancel</a>
                    </div>
                </form>
            </div>
        </div>
    </div>
</div>
''',

    'hall_list': '''
<div class="page-hero">
    <div class="page-title-block">
        <p class="page-kicker">Venues</p>
        <h2 class="mb-1">Halls & Venues</h2>
        <p class="text-muted mb-0">Manage the spaces available for events and keep capacity details easy to review.</p>
    </div>
    <div class="page-toolbar">
        <a href="/halls/add" class="btn btn-primary">+ Add Hall</a>
    </div>
</div>
<div class="card">
    <div class="card-body">
        <div class="table-responsive">
            <table class="table table-hover">
                <thead>
                    <tr><th>Hall Name</th><th>Capacity</th><th>Events Scheduled</th><th>Actions</th></tr>
                </thead>
                <tbody>
                    {hall_rows}
                </tbody>
            </table>
        </div>
    </div>
</div>
''',

    'hall_form': '''
<div class="page-hero">
    <div class="page-title-block">
        <p class="page-kicker">Venues</p>
        <h2 class="mb-1">{title}</h2>
        <p class="text-muted mb-0">Capture venue details clearly so every event can be assigned with confidence.</p>
    </div>
</div>
<div class="row justify-content-center">
    <div class="col-lg-6">
        <div class="card">
            <div class="card-body">
                <form method="POST">
                    <div class="mb-3">
                        <label class="form-label">Hall Name *</label>
                        <input type="text" name="hall_name" class="form-control" required value="{hall_name}">
                    </div>
                    <div class="mb-3">
                        <label class="form-label">Capacity *</label>
                        <input type="number" name="capacity" class="form-control" min="1" required value="{capacity}">
                    </div>
                    <div class="d-flex gap-2">
                        <button type="submit" class="btn btn-primary">Save</button>
                        <a href="/halls" class="btn btn-secondary">Cancel</a>
                    </div>
                </form>
            </div>
        </div>
    </div>
</div>
'''
}


class EventSchedulerHandler(http.server.SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=os.path.dirname(__file__) or '.', **kwargs)

    def check_auth(self, headers):
        cookie = headers.get('Cookie', '')
        session_cookie = None
        for c in cookie.split(';'):
            c = c.strip()
            if c.startswith('session='):
                session_cookie = c[8:]
                break
        return session_cookie in VALID_CREDENTIALS.values()

    def require_auth(self):
        if not self.check_auth(self.headers):
            self.send_response(302)
            self.send_header('Location', '/login')
            self.end_headers()
            return False
        return True

    def render_template(self, template_name, **kwargs):
        base = TEMPLATES['base.html']
        content = TEMPLATES.get(template_name, '').format(**kwargs)
        flash_msg = kwargs.get('flash', '')
        if flash_msg:
            flash_html = f'<div class="alert alert-success alert-dismissible fade show" role="alert">{flash_msg}<button type="button" class="btn-close" data-bs-dismiss="alert"></button></div>'
        else:
            flash_html = ''
        return base.format(
            title=kwargs.get('title', 'Event Scheduler'),
            flash=flash_html,
            content=content
        )

    def render_login(self, error_message=''):
        return LOGIN_TEMPLATE.format(error_message=error_message)

    def send_html(self, html, status=200):
        self.send_response(status)
        self.send_header('Content-Type', 'text/html; charset=utf-8')
        self.end_headers()
        self.wfile.write(html.encode('utf-8'))

    def send_json(self, data, status=200):
        self.send_response(status)
        self.send_header('Content-Type', 'application/json')
        self.end_headers()
        self.wfile.write(json.dumps(data).encode('utf-8'))

    def send_redirect(self, url, flash=None):
        self.send_response(302)
        self.send_header('Location', url)
        if flash:
            self.send_header('Set-Cookie', f'flash={escape_html(flash)}; Path=/')
        self.end_headers()

    def do_GET(self):
        parsed = urlparse(self.path)
        path = parsed.path
        query = parse_qs(parsed.query)

        if path == '/login':
            self.handle_login_get()
            return
        if path == '/logout':
            self.handle_logout()
            return
        if path.startswith('/static/'):
            super().do_GET()
            return
        if not self.require_auth():
            return

        if path == '/' or path == '/dashboard':
            self.handle_dashboard(query)
        elif path == '/create-event':
            self.handle_create_event_get(query)
        elif path == '/events':
            self.handle_event_list(query)
        elif path == '/upload-excel':
            self.handle_upload_excel_get()
        elif re.match(r'^/event/(\d+)$', path):
            event_id = int(re.match(r'^/event/(\d+)$', path).group(1))
            self.handle_event_detail(event_id)
        elif re.match(r'^/event/(\d+)/edit$', path):
            event_id = int(re.match(r'^/event/(\d+)/edit$', path).group(1))
            self.handle_edit_event_get(event_id)
        elif path == '/faculty':
            self.handle_faculty_list(query)
        elif path == '/faculty/add':
            self.handle_faculty_form()
        elif re.match(r'^/faculty/(\d+)/edit$', path):
            faculty_id = int(re.match(r'^/faculty/(\d+)/edit$', path).group(1))
            self.handle_faculty_form(faculty_id)
        elif re.match(r'^/faculty/(\d+)/availability$', path):
            faculty_id = int(re.match(r'^/faculty/(\d+)/availability$', path).group(1))
            self.handle_faculty_availability(faculty_id)
        elif path == '/halls':
            self.handle_hall_list()
        elif path == '/halls/add':
            self.handle_hall_form()
        elif re.match(r'^/halls/(\d+)/edit$', path):
            hall_id = int(re.match(r'^/halls/(\d+)/edit$', path).group(1))
            self.handle_hall_form(hall_id)
        elif path.startswith('/api/'):
            self.handle_api_get(path, query)
        else:
            self.send_error(404)

    def do_POST(self):
        parsed = urlparse(self.path)
        path = parsed.path

        if path == '/login':
            content_length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(content_length).decode('utf-8')
            self.handle_login_post(body)
            return

        if not self.require_auth():
            return

        content_length = int(self.headers.get('Content-Length', 0))
        
        # Check if this is a file upload
        content_type = self.headers.get('Content-Type', '')
        if 'multipart/form-data' in content_type:
            # For file uploads, we handle the raw data differently
            if path == '/upload-excel':
                self.handle_upload_excel_post()
            else:
                self.send_error(404)
            return
        
        # For regular POST requests, decode as UTF-8
        body = self.rfile.read(content_length).decode('utf-8')

        if path == '/create-event':
            self.handle_create_event_post(body)
        elif re.match(r'^/event/(\d+)/edit$', path):
            event_id = int(re.match(r'^/event/(\d+)/edit$', path).group(1))
            self.handle_edit_event_post(event_id, body)
        elif re.match(r'^/event/(\d+)/delete$', path):
            event_id = int(re.match(r'^/event/(\d+)/delete$', path).group(1))
            self.handle_delete_event(event_id)
        elif path == '/faculty/add':
            self.handle_faculty_post(body)
        elif re.match(r'^/faculty/(\d+)/edit$', path):
            faculty_id = int(re.match(r'^/faculty/(\d+)/edit$', path).group(1))
            self.handle_faculty_post(body, faculty_id)
        elif re.match(r'^/faculty/(\d+)/delete$', path):
            faculty_id = int(re.match(r'^/faculty/(\d+)/delete$', path).group(1))
            self.handle_delete_faculty(faculty_id)
        elif re.match(r'^/faculty/(\d+)/availability$', path):
            faculty_id = int(re.match(r'^/faculty/(\d+)/availability$', path).group(1))
            self.handle_faculty_availability_post(faculty_id, body)
        elif path == '/halls/add':
            self.handle_hall_post(body)
        elif re.match(r'^/halls/(\d+)/edit$', path):
            hall_id = int(re.match(r'^/halls/(\d+)/edit$', path).group(1))
            self.handle_hall_post(body, hall_id)
        elif re.match(r'^/halls/(\d+)/delete$', path):
            hall_id = int(re.match(r'^/halls/(\d+)/delete$', path).group(1))
            self.handle_delete_hall(hall_id)
        elif path == '/check-conflicts':
            self.handle_check_conflicts(body)
        elif path == '/send-reminders':
            self.handle_send_reminders()
        else:
            self.send_error(404)

    def do_PUT(self):
        parsed = urlparse(self.path)
        path = parsed.path
        if not self.require_auth():
            return
        content_length = int(self.headers.get('Content-Length', 0))
        body = self.rfile.read(content_length).decode('utf-8')
        if re.match(r'^/event/(\d+)/edit$', path):
            event_id = int(re.match(r'^/event/(\d+)/edit$', path).group(1))
            self.handle_edit_event_post(event_id, body)
        else:
            self.send_error(404)

    def handle_login_get(self):
        if self.check_auth(self.headers):
            self.send_redirect('/')
            return
        html = self.render_login()
        self.send_html(html)

    def handle_login_post(self, body):
        data = parse_form(body)
        username = data.get('username', '').strip()
        password = data.get('password', '').strip()
        if username in VALID_CREDENTIALS and VALID_CREDENTIALS[username] == password:
            self.send_response(302)
            self.send_header('Location', '/')
            self.send_header('Set-Cookie', f'session={password}; Path=/')
            self.end_headers()
        else:
            error_msg = '<div class="alert alert-danger">Invalid username or password</div>'
            html = self.render_login(error_message=error_msg)
            self.send_html(html, status=401)

    def handle_logout(self):
        self.send_response(302)
        self.send_header('Location', '/login')
        self.send_header('Set-Cookie', 'session=; Path=/; Expires=Thu, 01 Jan 1970 00:00:00 GMT')
        self.end_headers()

    def handle_upload_excel_get(self):
        html = self.render_template('upload_excel', title='Import Excel')
        self.send_html(html)

    def handle_upload_excel_post(self):
        try:
            # Parse the multipart form data
            form = cgi.FieldStorage(
                fp=self.rfile,
                headers=self.headers,
                environ={'REQUEST_METHOD': 'POST', 'CONTENT_TYPE': self.headers.get('Content-Type', '')}
            )

            if 'excelFile' not in form:
                self.send_json({'success': False, 'message': 'No file uploaded'}, 400)
                return

            file_item = form['excelFile']
            if not file_item.filename:
                self.send_json({'success': False, 'message': 'No file selected'}, 400)
                return

            clear_existing = 'clearExisting' in form and form['clearExisting'].value == 'on'

            # Save the file temporarily
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            temp_file.write(file_item.file.read())
            temp_file.close()

            try:
                # Process the Excel file
                result = self.process_excel_file(temp_file.name, clear_existing)
                os.unlink(temp_file.name)
                
                if result['success']:
                    self.send_redirect('/events', f'Successfully imported {result["created"]} events')
                else:
                    self.send_json(result, 400)
            except Exception as e:
                os.unlink(temp_file.name)
                self.send_json({'success': False, 'message': f'Error processing file: {str(e)}'}, 500)
        except Exception as e:
            self.send_json({'success': False, 'message': f'Error: {str(e)}'}, 500)

    def _parse_sheet_date(self, sheet, sheet_name):
        for row in range(1, min(5, sheet.max_row + 1)):
            cell_val = sheet.cell(row, 1).value
            if isinstance(cell_val, datetime):
                return cell_val.strftime('%Y-%m-%d')
            if isinstance(cell_val, str):
                date_match = re.search(r'(\d{1,2})\.(\d{1,2})\.(\d{4})', cell_val)
                if date_match:
                    day, month, year = date_match.groups()
                    return f'{year}-{month.zfill(2)}-{day.zfill(2)}'
        sheet_date_match = re.match(r'^(\d{1,2})\.(\d{1,2})$', sheet_name)
        if sheet_date_match:
            day, month = sheet_date_match.groups()
            return f'2026-{month.zfill(2)}-{day.zfill(2)}'
        return None

    def _normalize_time_slot(self, raw_value):
        if not raw_value:
            return None
        if isinstance(raw_value, str):
            value = raw_value.strip()
        else:
            value = str(raw_value).strip()
        if not value:
            return None
        if re.match(r'^\d{1,2}:\d{2}-\d{1,2}:\d{2}$', value):
            return value
        return None

    def process_excel_file(self, file_path, clear_existing):
        """Process the Excel file and create events based on the uploaded workbook structure."""
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            return {'success': False, 'message': f'Could not read Excel file: {str(e)}'}

        db = get_db()
        created_count = 0

        if clear_existing:
            db.execute('DELETE FROM event_faculty')
            db.execute('DELETE FROM event_branch_section')
            db.execute('DELETE FROM event')
            db.commit()

        time_slot_map = {
            '08:30-09:30': 1,
            '09:45-10:45': 2,
            '11:00-12:00': 3,
            '02:00-03:00': 4,
            '03:15-04:15': 5,
        }

        for sheet_name in wb.sheetnames:
            if sheet_name in ['Index', 'Core Team Duties', 'GA Schedule', 'Check list']:
                continue

            sheet = wb[sheet_name]
            date_str = self._parse_sheet_date(sheet, sheet_name)
            if not date_str:
                continue

            section_headers = []
            for col in range(2, sheet.max_column + 1):
                cell_val = sheet.cell(2, col).value
                if cell_val and isinstance(cell_val, str):
                    section_name = cell_val.strip()
                    if section_name and not section_name.startswith('Event') and not section_name.startswith('Sl'):
                        section_headers.append((col, section_name))

            if not section_headers:
                continue

            for row in range(3, sheet.max_row + 1):
                time_cell = sheet.cell(row, 1).value
                time_str = self._normalize_time_slot(time_cell)
                if not time_str:
                    continue
                if time_str not in time_slot_map:
                    continue

                period_no = time_slot_map[time_str]
                event_row_data = [self._normalize_excel_text(sheet.cell(row, col).value) for col in range(1, sheet.max_column + 1)]
                venue_row_data = [self._normalize_excel_text(sheet.cell(row + 1, col).value) for col in range(1, sheet.max_column + 1)] if row + 1 <= sheet.max_row else []
                faculty_row_data = [self._normalize_excel_text(sheet.cell(row + 2, col).value) for col in range(1, sheet.max_column + 1)] if row + 2 <= sheet.max_row else []

                for col, section_name in section_headers:
                    if col >= len(event_row_data):
                        continue
                    event_name = event_row_data[col]
                    if not event_name or event_name.lower() in {'event', 'venue', 'faculty assigned', 'faculty', 'inauguration'}:
                        continue
                    venue = venue_row_data[col] if col < len(venue_row_data) else ''
                    faculty = faculty_row_data[col] if col < len(faculty_row_data) else ''
                    self.create_event_from_excel(db, date_str, period_no, event_name, section_name, venue, faculty)
                    created_count += 1

        db.close()
        return {'success': True, 'created': created_count}

    def create_event_from_excel(self, db, date_str, period_no, event_name, section_name, venue, faculty):
        """Create a single event from Excel data"""
        try:
            event_name = self._normalize_event_name(event_name)
            venue = self._normalize_venue_name(venue)
            
            if not event_name:
                return

            # Parse section name (e.g., "CSE A" -> branch="CSE", section="A")
            parts = section_name.strip().split()
            if len(parts) >= 2:
                branch_name = parts[0]
                section_letter = parts[1]
            else:
                branch_name = section_name
                section_letter = 'A'

            # Get or create branch
            branch = db.execute('SELECT id FROM branch WHERE LOWER(name) = LOWER(?)', (branch_name,)).fetchone()
            if not branch:
                cursor = db.execute('INSERT INTO branch (name) VALUES (?)', (branch_name,))
                branch_id = cursor.lastrowid
            else:
                branch_id = branch['id']

            # Get or create section
            section = db.execute('SELECT id FROM section WHERE branch_id = ? AND LOWER(section_name) = LOWER(?)',
                                (branch_id, section_letter)).fetchone()
            if not section:
                cursor = db.execute('INSERT INTO section (branch_id, section_name) VALUES (?, ?)',
                                   (branch_id, section_letter))
                section_id = cursor.lastrowid
            else:
                section_id = section['id']

            # Get time slot
            time_slot = db.execute('SELECT id FROM time_slot WHERE period_no = ?', (period_no,)).fetchone()
            if not time_slot:
                return
            timeslot_id = time_slot['id']

            # Get or create hall based on venue text
            hall_id = None
            if venue:
                hall = db.execute('SELECT id FROM hall WHERE LOWER(hall_name) = LOWER(?)', (venue,)).fetchone()
                if hall:
                    hall_id = hall['id']
                else:
                    cursor = db.execute('INSERT INTO hall (hall_name, capacity) VALUES (?, ?)', (venue, 100))
                    hall_id = cursor.lastrowid
            if not hall_id:
                default_hall = db.execute('SELECT id FROM hall LIMIT 1').fetchone()
                if default_hall:
                    hall_id = default_hall['id']
                else:
                    return

            # Avoid importing duplicates for same event and section
            existing = db.execute('''
                SELECT e.id FROM event e
                JOIN event_branch_section ebs ON e.id = ebs.event_id
                WHERE e.date = ? AND e.timeslot_id = ? 
                AND ebs.branch_id = ? AND ebs.section_id = ?
                AND e.event_name = ?
            ''', (date_str, timeslot_id, branch_id, section_id, event_name)).fetchone()
            if existing:
                return

            # Determine event type
            event_type = self._infer_event_type_from_name(event_name)

            faculty_names = self._split_faculty_names(faculty)
            conducted_by = ', '.join(faculty_names) if faculty_names else ''
            remarks = f"Imported from Excel"
            if venue:
                remarks += f" | Venue: {venue}"
            if faculty:
                remarks += f" | Faculty: {faculty}"

            cursor = db.execute('''
                INSERT INTO event (event_name, event_type, description, conducted_by, date, timeslot_id, hall_id, remarks)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (event_name, event_type, f"Imported from Excel: {event_name}", conducted_by, date_str, timeslot_id, hall_id, remarks))

            event_id = cursor.lastrowid

            # Link event to branch and section
            db.execute('''
                INSERT INTO event_branch_section (event_id, branch_id, section_id)
                VALUES (?, ?, ?)
            ''', (event_id, branch_id, section_id))

            # Link event to faculty members if names can be found or created
            for faculty_name in faculty_names:
                faculty_id = self._find_or_create_faculty(db, faculty_name)
                if faculty_id:
                    db.execute('INSERT OR IGNORE INTO event_faculty (event_id, faculty_id) VALUES (?, ?)',
                               (event_id, faculty_id))

            db.commit()
        except Exception as e:
            print(f"Error creating event: {e}")
            db.rollback()

    def _normalize_excel_text(self, value):
        if value is None:
            return ''
        return str(value).strip()

    def _normalize_venue_name(self, venue_text):
        if not venue_text:
            return ''
        v = str(venue_text).strip().lower()
        # Remove extra spaces
        v = re.sub(r'\s+', ' ', v)
        
        # Check for synonyms/abbreviations
        if v in {'ah', 'amriteswari', 'amriteshwari', 'amriteshwari hall', 'amriteswari hall', 'amriteshari', 'amriteshari hall'}:
            return 'Amriteshwari Hall'
        if v in {'sh', 'sudhamani', 'sudhamani hall'}:
            return 'Sudhamani Hall'
        if v in {'eh', 'e-learning hall', 'e-learning'}:
            return 'E-Learning Hall'
        if v in {'cp-1', 'cp lab 1', 'cp lab -1', 'cp lab - 1', 'comp. lab. 1'}:
            return 'CP Lab - 1'
        if v in {'cp-2', 'cp lab 2', 'cp lab -2', 'cp lab - 2'}:
            return 'CP Lab - 2'
        if v in {'cp-3', 'cp lab 3', 'cp lab -3', 'cp lab - 3', 'cp lab - iii'}:
            return 'CP Lab - 3'
        if v in {'cp-4', 'cp lab 4', 'cp lab -4', 'cp lab - 4'}:
            return 'CP Lab - 4'
        if v in {'classroom', 'class room'}:
            return 'Classroom'
        if v in {'simulation lab', 'simulation', 'simulationlab'}:
            return 'Simulation Lab'
        if v in {'yoga hall', 'yoga'}:
            return 'Yoga Hall'
        if v in {'inauguration', 'event', 'venue', 'faculty assigned', 'faculty'}:
            return ''
            
        # Default capitalization
        words = []
        for w in str(venue_text).strip().split():
            if w.lower() in {'lab', 'hall', 'classroom', 'ground'}:
                words.append(w.capitalize())
            elif w.upper() in {'CP', 'AI', 'ECE', 'CAE'}:
                words.append(w.upper())
            else:
                words.append(w.capitalize())
        return ' '.join(words)

    def _normalize_event_name(self, event_name):
        if not event_name:
            return ''
        name = str(event_name).replace('\xa0', ' ').replace('\n', ' ').strip()
        # Remove extra spaces
        name = re.sub(r'\s+', ' ', name)
        name_lower = name.lower()
        
        # Check placeholders
        if name_lower in {'event', 'venue', 'faculty assigned', 'faculty', 'sl. no.'}:
            return ''
            
        # Standardize Bridge Course names (BrCo X)
        brco_match = re.search(r'\b(?:br\s*co|bridge\s*course)\s*[-]*\s*(\d+)\b', name_lower)
        if brco_match:
            return f"BrCo {brco_match.group(1)}"
            
        # Standardize Special Events names (SE X)
        se_match = re.search(r'\bse\s*[-]*\s*(\d+)\b', name_lower)
        if se_match:
            return f"SE {se_match.group(1)}"
            
        # Standardize General Addresses
        if name_lower in {'general addresses i', 'ga 1', 'ga1', 'general address i'}:
            return 'GA 1'
        if name_lower in {'general addresses ii', 'ga 2', 'ga2', 'general address ii'}:
            return 'GA 2'
            
        # Standardize Alumni sessions
        if 'alumni' in name_lower or 'alumini' in name_lower:
            # If it has speaker name in parentheses, normalize: "Alumni Session (Mr. Amit Kr. Pandey)"
            speaker_match = re.search(r'\(([^)]+)\)', name)
            if speaker_match:
                speaker = speaker_match.group(1).strip()
                # Normalize prefix inside parentheses
                speaker = re.sub(r'^Dr\.?\s*', 'Dr. ', speaker, flags=re.IGNORECASE)
                speaker = re.sub(r'^Mr\.?\s*', 'Mr. ', speaker, flags=re.IGNORECASE)
                speaker = re.sub(r'^Ms\.?\s*', 'Ms. ', speaker, flags=re.IGNORECASE)
                return f"Alumni Session ({speaker})"
            return 'Alumni Session'
            
        # Standardize Counselling
        if name_lower in {'counselling', 'counselling (cg)', 'counselling cg'}:
            return 'Counselling (CG)'
            
        # Standardize Department Session
        if name_lower in {'dept', 'dept.', 'dept. session', 'department session', 'dept. prof.'}:
            return 'Dept. Session'
            
        return name

    def _normalize_faculty_name(self, name):
        # Replace non-breaking spaces and clean
        name = str(name).replace('\xa0', ' ').strip()
        # Remove trailing/leading punctuation
        name = re.sub(r'^\s*[,;&/\s]+', '', name)
        name = re.sub(r'[,;&/\s]+\s*$', '', name)
        name = name.strip()
        
        name_lower = name.lower()
        if not name or name_lower in {'inauguration', 'fas', 'fa', 'no activity', 'tbd', 'tba', 'dept', 'dept.', 'department', 'faculty', 'faculty assigned', 'committee', 'coordinators', 'ieee', 'disha bharat'}:
            return None
            
        # Normalize Dr/Mr/Ms prefix (ensure they have a period and one space)
        name = re.sub(r'^Dr\.?\s*', 'Dr. ', name, flags=re.IGNORECASE)
        name = re.sub(r'^Mr\.?\s*', 'Mr. ', name, flags=re.IGNORECASE)
        name = re.sub(r'^Ms\.?\s*', 'Ms. ', name, flags=re.IGNORECASE)
        name = re.sub(r'\s+', ' ', name) # collapse extra spaces
        
        name_lower = name.lower()
        # Standardize specific duplicate names we found
        if 'aiswariya' in name_lower:
            return 'Ms. Aiswariya Milan K'
        if 'kruthika' in name_lower:
            return 'Ms. U. Kruthika'
        if 'deepthi janardhan' in name_lower:
            return 'Dr. Deepthi Janardhan'
        if 'sayanth vijay' in name_lower:
            return 'Dr. Sayanth Vijay'
        if 'phani raj' in name_lower:
            return 'Dr. Phani Raj Harivanam'
        if 'kirthika devi' in name_lower and 'support' not in name_lower:
            return 'Dr. Kirthika Devi V. S.'
        if 'amrita thakur' in name_lower:
            return 'Dr. Amrita Thakur'
        if 'manitha' in name_lower:
            return 'Dr. Manitha P. V.'
            
        return name

    def _split_faculty_names(self, faculty_text):
        if not faculty_text:
            return []
        
        # Replace non-breaking spaces and standardize
        text = str(faculty_text).replace('\xa0', ' ')
        # Split by delimiters: commas, slashes, ampersands, semicolons, "and", and "Support-" variations
        delims = [r'\s*,\s*', r'\s*/\s*', r'\s*&\s*', r'\s+and\s+', r'\s*;\s*', r'\s*\n\s*', r'\s+Support-\s*', r'\s+Support\s+-\s*', r'\s+Support\s+', r'\s+support\s+']
        combined_delim = '|'.join(delims)
        
        normalized_text = re.sub(combined_delim, '|', text, flags=re.IGNORECASE)
        parts = normalized_text.split('|')
        
        results = []
        for p in parts:
            p_clean = self._normalize_faculty_name(p)
            if p_clean:
                results.append(p_clean)
        return results

    def _find_or_create_faculty(self, db, faculty_name):
        faculty_name = faculty_name.strip()
        if not faculty_name:
            return None
        faculty = db.execute('SELECT id FROM faculty WHERE LOWER(TRIM(faculty_name)) = LOWER(?)', (faculty_name,)).fetchone()
        if faculty:
            return faculty['id']
        cursor = db.execute('INSERT INTO faculty (faculty_name, department, email, phone) VALUES (?, ?, ?, ?)',
                            (faculty_name, 'Imported', None, None))
        return cursor.lastrowid

    def _infer_event_type_from_name(self, event_name):
        name_lower = event_name.lower()
        if 'inauguration' in name_lower:
            return 'Inauguration'
        if 'counselling' in name_lower or 'cg' in name_lower:
            return 'Counselling'
        if 'sports' in name_lower:
            return 'Sports'
        if 'alumni' in name_lower or 'alumini' in name_lower:
            return 'Alumni Session'
        if 'brco' in name_lower or 'bridge course' in name_lower or 'br co' in name_lower:
            return 'Bridge Course'
        if 'dpe' in name_lower or 'dastaan' in name_lower:
            return 'Dastaan Pre-Event'
        if 'ga 1' in name_lower or 'ga 2' in name_lower or 'general address' in name_lower:
            return 'General Address'
        if 'disha bharat' in name_lower:
            return 'Disha Bharat'
        if 'meditation' in name_lower:
            return 'Meditation'
        if 'uhv' in name_lower:
            return 'UHV'
        if 'dept' in name_lower or 'department' in name_lower:
            return 'Dept. Session'
        if 'icc' in name_lower:
            return 'ICC'
        if 'se ' in name_lower or 'special event' in name_lower or name_lower.startswith('se'):
            return 'Special Event'
        return 'Others'

    def _to_int(self, value, field_name, required=True):
        if value in (None, ''):
            if required:
                raise ValueError(f'{field_name} is required')
            return None
        try:
            return int(value)
        except (TypeError, ValueError):
            raise ValueError(f'{field_name} must be a valid number')

    def _parse_event_payload(self, data):
        required_text = {
            'event_name': 'Event name',
            'event_type': 'Event type',
            'conducted_by': 'Conducted by',
            'date': 'Event date',
        }
        payload = {}
        for key, label in required_text.items():
            value = str(data.get(key, '')).strip()
            if not value:
                raise ValueError(f'{label} is required')
            payload[key] = value

        try:
            datetime.strptime(payload['date'], '%Y-%m-%d')
        except ValueError:
            raise ValueError('Event date must use YYYY-MM-DD format')

        payload['description'] = str(data.get('description', '') or '').strip()
        payload['remarks'] = str(data.get('remarks', '') or '').strip()
        payload['timeslot_id'] = self._to_int(data.get('timeslot_id'), 'Time slot')
        payload['hall_id'] = self._to_int(data.get('hall_id'), 'Hall')
        payload['participants'] = self._to_int(data.get('participants'), 'Participants', required=False)
        if payload['participants'] is not None and payload['participants'] < 1:
            raise ValueError('Participants must be at least 1')

        sections = data.get('sections', []) or []
        if not isinstance(sections, list):
            raise ValueError('Sections must be a list')
        payload['sections'] = []
        seen_sections = set()
        for section in sections:
            if not isinstance(section, dict):
                continue
            branch_id = section.get('branchId')
            section_id = section.get('sectionId')
            if branch_id and section_id:
                key = f"{branch_id}_{section_id}"
                if key not in seen_sections:
                    payload['sections'].append({'branch_id': branch_id, 'section_id': section_id})
                    seen_sections.add(key)

        if not payload['sections']:
            raise ValueError('At least one section must be selected')

        faculty_ids = data.get('faculty_ids', []) or []
        if not isinstance(faculty_ids, list):
            raise ValueError('Faculty IDs must be a list')
        payload['faculty_ids'] = []
        seen_faculty = set()
        for faculty_id in faculty_ids:
            parsed_id = self._to_int(faculty_id, 'Faculty', required=False)
            if parsed_id and parsed_id not in seen_faculty:
                payload['faculty_ids'].append(parsed_id)
                seen_faculty.add(parsed_id)

        return payload

    def _validate_event_references(self, db, payload, event_id=None):
        if event_id and not db.execute('SELECT id FROM event WHERE id = ?', (event_id,)).fetchone():
            raise ValueError('Event not found')

        for section in payload['sections']:
            section_record = db.execute(
                'SELECT id FROM section WHERE id = ? AND branch_id = ?',
                (section['section_id'], section['branch_id'])
            ).fetchone()
            if not section_record:
                raise ValueError(f"Section {section['section_id']} does not belong to branch {section['branch_id']}")

        if not db.execute('SELECT id FROM time_slot WHERE id = ?', (payload['timeslot_id'],)).fetchone():
            raise ValueError('Time slot not found')

        hall = db.execute('SELECT id, capacity FROM hall WHERE id = ?', (payload['hall_id'],)).fetchone()
        if not hall:
            raise ValueError('Hall not found')
        if payload['participants'] and payload['participants'] > hall['capacity']:
            raise ValueError('Expected participants exceed selected hall capacity')

        if payload['faculty_ids']:
            placeholders = ','.join('?' for _ in payload['faculty_ids'])
            found = {
                row['id'] for row in
                db.execute(f'SELECT id FROM faculty WHERE id IN ({placeholders})', payload['faculty_ids']).fetchall()
            }
            missing = [str(fid) for fid in payload['faculty_ids'] if fid not in found]
            if missing:
                raise ValueError(f'Faculty not found: {", ".join(missing)}')

    def _find_conflicts(self, db, payload, event_id=None):
        event_date = payload['date']
        timeslot_id = payload['timeslot_id']
        hall_id = payload['hall_id']
        faculty_ids = payload.get('faculty_ids', [])
        sections = payload.get('sections', [])
        conflicts = []

        event_day = datetime.strptime(event_date, '%Y-%m-%d').weekday()

        for section in sections:
            slot_query = '''
                SELECT e.id FROM event e
                JOIN event_branch_section ebs ON e.id = ebs.event_id
                WHERE e.date = ? AND e.timeslot_id = ? 
                AND ebs.branch_id = ? AND ebs.section_id = ?
            '''
            params = [event_date, timeslot_id, section['branch_id'], section['section_id']]
            if event_id:
                slot_query += ' AND e.id != ?'
                params.append(event_id)

            if db.execute(slot_query, params).fetchone():
                branch = db.execute('SELECT name FROM branch WHERE id = ?', (section['branch_id'],)).fetchone()
                section_record = db.execute('SELECT section_name FROM section WHERE id = ?', (section['section_id'],)).fetchone()
                conflicts.append({
                    'type': 'branch',
                    'message': f"{branch['name']}-{section_record['section_name']} already has an event at this time slot."
                })
                break

        faculty_conflicts = []
        for faculty_id in faculty_ids:
            faculty = db.execute('SELECT faculty_name FROM faculty WHERE id = ?', (faculty_id,)).fetchone()
            faculty_name = faculty['faculty_name'] if faculty else f"ID {faculty_id}"
            
            availability = db.execute('''
                SELECT is_available FROM faculty_availability
                WHERE faculty_id = ? AND day_of_week = ? AND timeslot_id = ? AND is_available = 0
            ''', (faculty_id, event_day, timeslot_id)).fetchone()
            
            if availability:
                faculty_conflicts.append(f"{faculty_name} (marked as unavailable on this day)")

            faculty_query = '''
                SELECT f.faculty_name FROM event_faculty ef
                JOIN event e ON ef.event_id = e.id
                JOIN faculty f ON ef.faculty_id = f.id
                WHERE e.date = ? AND e.timeslot_id = ? AND ef.faculty_id = ?
            '''
            params = [event_date, timeslot_id, faculty_id]
            if event_id:
                faculty_query += ' AND e.id != ?'
                params.append(event_id)

            existing_event = db.execute(faculty_query, params).fetchone()
            if existing_event:
                faculty_conflicts.append(f"{existing_event['faculty_name']} (already assigned to another event)")

        if faculty_conflicts:
            conflicts.append({
                'type': 'faculty',
                'message': f"Faculty conflicts: {', '.join(faculty_conflicts)}",
                'alternatives': self._get_available_faculty_list(db, event_date, timeslot_id, event_id)
            })

        if hall_id:
            hall_query = 'SELECT id FROM event WHERE date = ? AND timeslot_id = ? AND hall_id = ?'
            params = [event_date, timeslot_id, hall_id]
            if event_id:
                hall_query += ' AND id != ?'
                params.append(event_id)

            if db.execute(hall_query, params).fetchone():
                hall = db.execute('SELECT hall_name FROM hall WHERE id = ?', (hall_id,)).fetchone()
                conflicts.append({
                    'type': 'hall',
                    'message': f"{hall['hall_name']} is already booked at this time.",
                    'alternatives': self._get_available_halls_list(db, event_date, timeslot_id, event_id)
                })

        return conflicts

    def _get_available_faculty_list(self, db, event_date, timeslot_id, exclude_event_id=None):
        event_day = datetime.strptime(event_date, '%Y-%m-%d').weekday()
        all_faculty = db.execute('SELECT id, faculty_name FROM faculty').fetchall()

        unavailable = set()
        availability = db.execute('''
            SELECT faculty_id FROM faculty_availability
            WHERE day_of_week = ? AND timeslot_id = ? AND is_available = 0
        ''', (event_day, timeslot_id)).fetchall()
        for a in availability:
            unavailable.add(a['faculty_id'])

        query = '''
            SELECT ef.faculty_id FROM event_faculty ef
            JOIN event e ON ef.event_id = e.id
            WHERE e.date = ? AND e.timeslot_id = ?
        '''
        params = [event_date, timeslot_id]
        if exclude_event_id:
            query += ' AND e.id != ?'
            params.append(exclude_event_id)

        busy_ids = set(row['faculty_id'] for row in db.execute(query, params).fetchall() if row['faculty_id'])

        available = []
        for f in all_faculty:
            if f['id'] not in unavailable and f['id'] not in busy_ids:
                available.append({'id': f['id'], 'name': f['faculty_name']})

        return available

    def _get_available_halls_list(self, db, event_date, timeslot_id, exclude_event_id=None):
        all_halls = db.execute('SELECT id, hall_name FROM hall').fetchall()

        query = 'SELECT hall_id FROM event WHERE date = ? AND timeslot_id = ?'
        params = [event_date, timeslot_id]
        if exclude_event_id:
            query += ' AND id != ?'
            params.append(exclude_event_id)

        busy_ids = set(row['hall_id'] for row in db.execute(query, params).fetchall())

        available = []
        for h in all_halls:
            if h['id'] not in busy_ids:
                available.append({'id': h['id'], 'name': h['hall_name']})

        return available

    def handle_dashboard(self, query=None):
        query = query or {}
        db = get_db()
        today = date.today().isoformat()
        date_from = normalize_date_filter(query.get('date_from', [today])[0], today)
        date_to = normalize_date_filter(query.get('date_to', [''])[0])
        if date_to and date_to < date_from:
            date_from, date_to = date_to, date_from
        branch_filter = normalize_int_filter(query.get('branch_id', [''])[0])
        type_filter = query.get('event_type', [''])[0]
        hall_filter = normalize_int_filter(query.get('hall_id', [''])[0])
        search_filter = query.get('q', [''])[0].strip()
        if branch_filter and not db.execute('SELECT id FROM branch WHERE id = ?', (branch_filter,)).fetchone():
            branch_filter = ''
        if hall_filter and not db.execute('SELECT id FROM hall WHERE id = ?', (hall_filter,)).fetchone():
            hall_filter = ''
        if type_filter and not db.execute('SELECT 1 FROM event WHERE event_type = ? LIMIT 1', (type_filter,)).fetchone():
            type_filter = ''

        today_count = db.execute('SELECT COUNT(*) FROM event WHERE date = ?', (today,)).fetchone()[0]
        faculty_count = db.execute('SELECT COUNT(*) FROM faculty').fetchone()[0]
        total_events = db.execute('SELECT COUNT(*) FROM event').fetchone()[0]
        hall_count = db.execute('SELECT COUNT(*) FROM hall').fetchone()[0]
        booked_halls_today = db.execute('SELECT COUNT(DISTINCT hall_id) FROM event WHERE date = ?', (today,)).fetchone()[0]
        available_halls_today = max(hall_count - booked_halls_today, 0)
        next7_count = db.execute(
            "SELECT COUNT(*) FROM event WHERE date BETWEEN ? AND date(?, '+6 days')",
            (today, today)
        ).fetchone()[0]

        filter_clauses = ['e.date >= ?']
        filter_params = [date_from]
        if date_to:
            filter_clauses.append('e.date <= ?')
            filter_params.append(date_to)
        if branch_filter:
            filter_clauses.append('EXISTS (SELECT 1 FROM event_branch_section ebs WHERE ebs.event_id = e.id AND ebs.branch_id = ?)')
            filter_params.append(branch_filter)
        if type_filter:
            filter_clauses.append('e.event_type = ?')
            filter_params.append(type_filter)
        if hall_filter:
            filter_clauses.append('e.hall_id = ?')
            filter_params.append(hall_filter)
        if search_filter:
            filter_clauses.append('(e.event_name LIKE ? OR e.conducted_by LIKE ? OR e.description LIKE ? OR e.remarks LIKE ?)')
            like = f'%{search_filter}%'
            filter_params.extend([like, like, like, like])
        filter_where = ' AND '.join(filter_clauses)

        upcoming = db.execute('''
            SELECT e.id, e.event_name, e.date, e.timeslot_id, e.hall_id,
                   ts.period_no, ts.start_time, ts.end_time,
                   h.hall_name, e.event_type, e.participants,
                   GROUP_CONCAT(DISTINCT b.name || '-' || s.section_name) as branch_sections,
                   GROUP_CONCAT(DISTINCT f.faculty_name) as faculty_names
            FROM event e
            JOIN time_slot ts ON e.timeslot_id = ts.id
            JOIN hall h ON e.hall_id = h.id
            LEFT JOIN event_branch_section ebs ON e.id = ebs.event_id
            LEFT JOIN branch b ON ebs.branch_id = b.id
            LEFT JOIN section s ON ebs.section_id = s.id
            LEFT JOIN event_faculty ef ON e.id = ef.event_id
            LEFT JOIN faculty f ON ef.faculty_id = f.id
            WHERE ''' + filter_where + '''
            GROUP BY e.id
            ORDER BY e.date, ts.period_no
            LIMIT 25
        ''', filter_params).fetchall()

        today_events = db.execute('''
            SELECT e.id, e.event_name, e.event_type,
                   ts.period_no, ts.start_time, ts.end_time, h.hall_name,
                   GROUP_CONCAT(DISTINCT b.name || '-' || s.section_name) as branch_sections,
                   GROUP_CONCAT(DISTINCT f.faculty_name) as faculty_names
            FROM event e
            JOIN time_slot ts ON e.timeslot_id = ts.id
            JOIN hall h ON e.hall_id = h.id
            LEFT JOIN event_branch_section ebs ON e.id = ebs.event_id
            LEFT JOIN branch b ON ebs.branch_id = b.id
            LEFT JOIN section s ON ebs.section_id = s.id
            LEFT JOIN event_faculty ef ON e.id = ef.event_id
            LEFT JOIN faculty f ON ef.faculty_id = f.id
            WHERE e.date = ?
            GROUP BY e.id
            ORDER BY ts.period_no
        ''', (today,)).fetchall()

        if upcoming:
            upcoming_table = '''<div class="table-responsive">
                <table class="table table-hover">
                    <thead><tr><th>Event</th><th>Date</th><th>Branches/Sections</th><th>Time</th><th>Hall</th><th>Type</th><th>Faculty</th></tr></thead>
                    <tbody>'''
            for e in upcoming:
                upcoming_table += f'''<tr>
                    <td><a href="/event/{e['id']}">{display_text(e['event_name'])}</a></td>
                    <td>{display_text(e['date'])}</td>
                    <td>{display_text(e['branch_sections'])}</td>
                    <td>Period {e['period_no']}: {display_text(e['start_time'])}-{display_text(e['end_time'])}</td>
                    <td>{display_text(e['hall_name'])}</td>
                    <td><span class="badge bg-secondary">{display_text(e['event_type'])}</span></td>
                    <td>{display_text(e['faculty_names'])}</td>
                </tr>'''
            upcoming_table += '</tbody></table></div>'
        else:
            upcoming_table = '<p class="text-muted mb-0">No events match the current filters</p>'

        if today_events:
            today_timeline = '<div class="timeline-list">'
            for e in today_events:
                today_timeline += f'''<a class="timeline-item" href="/event/{e['id']}">
                    <span class="timeline-time">P{e['period_no']}<small>{display_text(e['start_time'])}</small></span>
                    <span class="timeline-main">
                        <strong>{display_text(e['event_name'])}</strong>
                        <small>{display_text(e['branch_sections'])} | {display_text(e['hall_name'])}</small>
                    </span>
                    <span class="timeline-type">{display_text(e['event_type'])}</span>
                </a>'''
            today_timeline += '</div>'
        else:
            today_timeline = '<p class="text-muted mb-0">No events scheduled today.</p>'

        branches = db.execute('SELECT id, name FROM branch ORDER BY name').fetchall()
        halls = db.execute('SELECT id, hall_name FROM hall ORDER BY hall_name').fetchall()
        event_types = db.execute('SELECT DISTINCT event_type FROM event ORDER BY event_type').fetchall()

        branch_options = '<option value="">All branches</option>'
        for b in branches:
            selected = 'selected' if str(b['id']) == str(branch_filter) else ''
            branch_options += f'<option value="{b["id"]}" {selected}>{display_text(b["name"], "")}</option>'

        hall_options = '<option value="">All halls</option>'
        for h in halls:
            selected = 'selected' if str(h['id']) == str(hall_filter) else ''
            hall_options += f'<option value="{h["id"]}" {selected}>{display_text(h["hall_name"], "")}</option>'

        type_options = '<option value="">All types</option>'
        for t in event_types:
            selected = 'selected' if t['event_type'] == type_filter else ''
            type_options += f'<option value="{escape_html(t["event_type"])}" {selected}>{display_text(t["event_type"], "")}</option>'

        events_by_date = {}
        all_events = db.execute('''
            SELECT e.date, e.event_name, 
                   GROUP_CONCAT(DISTINCT b.name || '-' || s.section_name) as branch_sections,
                   GROUP_CONCAT(DISTINCT f.faculty_name) as faculty_names
            FROM event e
            LEFT JOIN event_branch_section ebs ON e.id = ebs.event_id
            LEFT JOIN branch b ON ebs.branch_id = b.id
            LEFT JOIN section s ON ebs.section_id = s.id
            LEFT JOIN event_faculty ef ON e.id = ef.event_id
            LEFT JOIN faculty f ON ef.faculty_id = f.id
            WHERE e.date >= ?
            GROUP BY e.id
        ''', (today,)).fetchall()

        for e in all_events:
            if e['date'] not in events_by_date:
                events_by_date[e['date']] = []
            events_by_date[e['date']].append({
                'name': e['event_name'],
                'branch_sections': e['branch_sections'] or 'N/A',
                'faculty': e['faculty_names'] or 'Not Assigned'
            })

        db.close()

        html = self.render_template('dashboard',
            title='Dashboard',
            today_count=today_count,
            faculty_count=faculty_count,
            total_events=total_events,
            next7_count=next7_count,
            available_halls_today=available_halls_today,
            hall_count=hall_count,
            upcoming_table=upcoming_table,
            today_timeline=today_timeline,
            branch_options=branch_options,
            hall_options=hall_options,
            type_options=type_options,
            date_from=escape_html(date_from),
            date_to=escape_html(date_to),
            search_filter=escape_html(search_filter),
            events_json=json.dumps(events_by_date))

        self.send_html(html)

    def handle_create_event_get(self, query):
        preselected_date = query.get('date', [''])[0]
        html = self.render_template('create_event',
            title='Create Event',
            preselected_date=preselected_date)
        self.send_html(html)

    def handle_create_event_post(self, body):
        try:
            payload = self._parse_event_payload(json.loads(body))
        except json.JSONDecodeError:
            self.send_json({'success': False, 'message': 'Invalid JSON'}, 400)
            return
        except ValueError as exc:
            self.send_json({'success': False, 'message': str(exc)}, 400)
            return

        db = get_db()
        try:
            self._validate_event_references(db, payload)
            conflicts = self._find_conflicts(db, payload)
            if conflicts:
                self.send_json({'success': False, 'message': 'Conflicts detected', 'conflicts': conflicts}, 409)
                return

            cursor = db.execute('''
                INSERT INTO event (event_name, event_type, description, conducted_by, date,
                                 timeslot_id, hall_id, participants, remarks)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                payload['event_name'],
                payload['event_type'],
                payload['description'],
                payload['conducted_by'],
                payload['date'],
                payload['timeslot_id'],
                payload['hall_id'],
                payload['participants'],
                payload['remarks']
            ))

            event_id = cursor.lastrowid

            for section in payload['sections']:
                db.execute('''
                    INSERT INTO event_branch_section (event_id, branch_id, section_id)
                    VALUES (?, ?, ?)
                ''', (event_id, section['branch_id'], section['section_id']))

            for faculty_id in payload['faculty_ids']:
                db.execute('INSERT INTO event_faculty (event_id, faculty_id) VALUES (?, ?)',
                          (event_id, faculty_id))
            db.commit()
        except ValueError as exc:
            db.rollback()
            self.send_json({'success': False, 'message': str(exc)}, 400)
            return
        except sqlite3.Error as exc:
            db.rollback()
            self.send_json({'success': False, 'message': f'Database error: {exc}'}, 500)
            return
        finally:
            db.close()

        self.send_json({'success': True, 'event_id': event_id})

    def handle_event_list(self, query=None):
        query = query or {}
        db = get_db()
 
        search_filter = (query.get('q', [''])[0] or '').strip()
        date_from = normalize_date_filter(query.get('date_from', [''])[0], '')
        date_to = normalize_date_filter(query.get('date_to', [''])[0], '')
        if date_to and date_from and date_to < date_from:
            date_from, date_to = date_to, date_from
        type_filter = (query.get('event_type', [''])[0] or '').strip()
        hall_filter = normalize_int_filter(query.get('hall_id', [''])[0])
        section_filter = normalize_int_filter(query.get('section_id', [''])[0])
 
        if hall_filter and not db.execute('SELECT id FROM hall WHERE id = ?', (hall_filter,)).fetchone():
            hall_filter = ''
        if type_filter and not db.execute('SELECT 1 FROM event WHERE event_type = ? LIMIT 1', (type_filter,)).fetchone():
            type_filter = ''
        if section_filter and not db.execute('SELECT id FROM section WHERE id = ?', (section_filter,)).fetchone():
            section_filter = ''
 
        clauses = []
        params = []
        if date_from:
            clauses.append('e.date >= ?')
            params.append(date_from)
        if date_to:
            clauses.append('e.date <= ?')
            params.append(date_to)
        if type_filter:
            clauses.append('e.event_type = ?')
            params.append(type_filter)
        if hall_filter:
            clauses.append('e.hall_id = ?')
            params.append(hall_filter)
        if section_filter:
            clauses.append('EXISTS (SELECT 1 FROM event_branch_section ebs WHERE ebs.event_id = e.id AND ebs.section_id = ?)')
            params.append(section_filter)
        if search_filter:
            clauses.append('(e.event_name LIKE ? OR e.conducted_by LIKE ? OR e.description LIKE ? OR e.remarks LIKE ? OR f.faculty_name LIKE ?)')
            like = f'%{search_filter}%'
            params.extend([like, like, like, like, like])
 
        where_clause = ''
        if clauses:
            where_clause = ' WHERE ' + ' AND '.join(clauses)
 
        events = db.execute(f'''
            SELECT e.*, ts.period_no, ts.start_time, ts.end_time, h.hall_name,
                   GROUP_CONCAT(DISTINCT b.name || '-' || s.section_name) as branch_sections,
                   GROUP_CONCAT(DISTINCT f.faculty_name) as faculty_names,
                   GROUP_CONCAT(DISTINCT s.id) as section_ids
            FROM event e
            JOIN time_slot ts ON e.timeslot_id = ts.id
            JOIN hall h ON e.hall_id = h.id
            LEFT JOIN event_branch_section ebs ON e.id = ebs.event_id
            LEFT JOIN branch b ON ebs.branch_id = b.id
            LEFT JOIN section s ON ebs.section_id = s.id
            LEFT JOIN event_faculty ef ON e.id = ef.event_id
            LEFT JOIN faculty f ON ef.faculty_id = f.id
            {where_clause}
            GROUP BY e.id
            ORDER BY e.date DESC, ts.period_no
        ''', params).fetchall()
 
        if events:
            events_table = f'''<div class="d-flex justify-content-between align-items-center mb-3">
                <span class="result-pill" id="resultCount">{len(events)} event(s) shown</span>
                <span class="text-muted small">Sorted by latest date first</span>
            </div>
            <div class="table-responsive">
                <table class="table table-hover" id="eventsTable">
                    <thead><tr><th>Event</th><th>Date</th><th>Branches/Sections</th><th>Time</th><th>Hall</th><th>Faculty</th><th>Type</th><th>Actions</th></tr></thead>
                    <tbody>'''
            for e in events:
                search_text = escape_html(f"{e['event_name']} {e['branch_sections'] or ''} {e['hall_name']} {e['faculty_names'] or ''} {e['event_type']}").lower()
                etype = str(e['event_type'] or '').lower()
                if 'inauguration' in etype:
                    b_cls = 'badge-inauguration'
                elif 'induction' in etype or 'orientation' in etype:
                    b_cls = 'badge-induction'
                elif 'workshop' in etype or 'seminar' in etype:
                    b_cls = 'badge-workshop'
                elif 'lecture' in etype or 'class' in etype:
                    b_cls = 'badge-lecture'
                elif 'meeting' in etype:
                    b_cls = 'badge-meeting'
                else:
                    b_cls = 'badge-other'
                events_table += f'''<tr data-date="{escape_html(e['date'])}" data-type="{escape_html(e['event_type'])}" data-hall-id="{escape_html(e['hall_id'])}" data-section-ids="{e['section_ids'] or ''}" data-search="{search_text}">
                    <td><a href="/event/{e['id']}">{display_text(e['event_name'])}</a></td>
                    <td>{display_text(e['date'])}</td>
                    <td>{display_text(e['branch_sections'])}</td>
                    <td>Period {e['period_no']}: {display_text(e['start_time'])}-{display_text(e['end_time'])}</td>
                    <td>{display_text(e['hall_name'])}</td>
                    <td>{display_text(e['faculty_names'])}</td>
                    <td><span class="badge {b_cls}">{display_text(e['event_type'])}</span></td>
                    <td>
                        <a href="/event/{e['id']}/edit" class="btn btn-sm btn-outline-primary">Edit</a>
                        <form action="/event/{e['id']}/delete" method="POST" style="display: inline;" onsubmit="return confirm('Delete this event?')">
                            <button type="submit" class="btn btn-sm btn-outline-danger">Delete</button>
                        </form>
                    </td>
                </tr>'''
            events_table += '</tbody></table></div>'
        else:
            events_table = '''<div class="empty-state">
                <p class="text-muted mb-3">No matching events were found.</p>
                <a href="/create-event" class="btn btn-primary">Create a New Event</a>
            </div>'''
 
        event_types = db.execute('SELECT DISTINCT event_type FROM event ORDER BY event_type').fetchall()
        halls = db.execute('SELECT id, hall_name FROM hall ORDER BY hall_name').fetchall()
        sections = db.execute('''
            SELECT s.id, s.section_name, b.name as branch_name 
            FROM section s 
            JOIN branch b ON s.branch_id = b.id 
            ORDER BY b.name, s.section_name
        ''').fetchall()
 
        type_options = '<option value="">All types</option>'
        for t in event_types:
            selected = 'selected' if t['event_type'] == type_filter else ''
            type_options += f'<option value="{escape_html(t["event_type"])}" {selected}>{display_text(t["event_type"], "")}</option>'
 
        hall_options = '<option value="">All venues</option>'
        for h in halls:
            selected = 'selected' if str(h['id']) == str(hall_filter) else ''
            hall_options += f'<option value="{h["id"]}" {selected}>{display_text(h["hall_name"], "")}</option>'
 
        section_options = '<option value="">All sections</option>'
        for s in sections:
            selected = 'selected' if str(s['id']) == str(section_filter) else ''
            section_options += f'<option value="{s["id"]}" {selected}>{display_text(s["branch_name"])}-{display_text(s["section_name"])}</option>'
 
        html = self.render_template('event_list',
            title='Events',
            events_table=events_table,
            search_filter=escape_html(search_filter),
            date_from=escape_html(date_from),
            date_to=escape_html(date_to),
            type_options=type_options,
            hall_options=hall_options,
            section_options=section_options)
        db.close()
        self.send_html(html)

    def handle_event_detail(self, event_id):
        db = get_db()
        event = db.execute('''
            SELECT e.*, ts.period_no, ts.start_time, ts.end_time,
                   h.hall_name, h.capacity as hall_capacity,
                   GROUP_CONCAT(DISTINCT b.name || '-' || s.section_name) as branch_sections,
                   GROUP_CONCAT(DISTINCT f.faculty_name || ' (' || f.department || ')') as faculty_names
            FROM event e
            JOIN time_slot ts ON e.timeslot_id = ts.id
            JOIN hall h ON e.hall_id = h.id
            LEFT JOIN event_branch_section ebs ON e.id = ebs.event_id
            LEFT JOIN branch b ON ebs.branch_id = b.id
            LEFT JOIN section s ON ebs.section_id = s.id
            LEFT JOIN event_faculty ef ON e.id = ef.event_id
            LEFT JOIN faculty f ON ef.faculty_id = f.id
            WHERE e.id = ?
            GROUP BY e.id
        ''', (event_id,)).fetchone()
        db.close()

        if not event:
            self.send_error(404)
            return

        html = self.render_template('event_detail',
            title=escape_html(event['event_name']),
            event_id=event_id,
            event_name=display_text(event['event_name'], ''),
            event_type=display_text(event['event_type'], ''),
            conducted_by=display_text(event['conducted_by']),
            participants=display_text(event['participants']),
            description=display_text(event['description'], 'No description provided'),
            remarks=display_text(event['remarks'], 'No remarks'),
            branches_sections=display_text(event['branch_sections'], 'N/A'),
            event_date=display_text(event['date'], ''),
            timeslot=f"Period {event['period_no']}: {display_text(event['start_time'])} - {display_text(event['end_time'])}",
            hall=f"{display_text(event['hall_name'])} (Capacity: {display_text(event['hall_capacity'])})",
            faculty=display_text(event['faculty_names'], 'Not assigned'))

        self.send_html(html)

    def handle_edit_event_get(self, event_id):
        db = get_db()
        
        event = db.execute('''
            SELECT e.*, ts.period_no, ts.start_time, ts.end_time
            FROM event e
            JOIN time_slot ts ON e.timeslot_id = ts.id
            WHERE e.id = ?
        ''', (event_id,)).fetchone()

        if not event:
            db.close()
            self.send_error(404)
            return

        current_sections = db.execute('''
            SELECT branch_id, section_id FROM event_branch_section WHERE event_id = ?
        ''', (event_id,)).fetchall()
        current_sections_list = [{'branchId': row['branch_id'], 'sectionId': row['section_id']} for row in current_sections]

        assigned_faculty = db.execute('''
            SELECT faculty_id FROM event_faculty WHERE event_id = ?
        ''', (event_id,)).fetchall()
        assigned_ids = [row['faculty_id'] for row in assigned_faculty]

        halls = db.execute('SELECT id, hall_name, capacity FROM hall ORDER BY hall_name').fetchall()
        time_slots = db.execute('SELECT * FROM time_slot ORDER BY period_no').fetchall()
        faculties = db.execute('SELECT id, faculty_name, department FROM faculty ORDER BY faculty_name').fetchall()
        db.close()

        slot_options = ''
        for slot in time_slots:
            selected = 'selected' if slot['id'] == event['timeslot_id'] else ''
            slot_options += f'<option value="{slot["id"]}" {selected}>Period {slot["period_no"]}: {display_text(slot["start_time"])} - {display_text(slot["end_time"])}</option>'

        type_options = ''
        for t in ['Workshop', 'Seminar', 'Technical Event', 'Club Event', 'Guest Lecture', 'Competition', 'Other']:
            selected = 'selected' if t == event['event_type'] else ''
            type_options += f'<option value="{escape_html(t)}" {selected}>{escape_html(t)}</option>'

        hall_options = ''
        for h in halls:
            selected = 'selected' if h['id'] == event['hall_id'] else ''
            hall_options += f'<option value="{h["id"]}" {selected}>{display_text(h["hall_name"], "")} (Capacity: {display_text(h["capacity"], "")})</option>'

        faculty_options = ''
        for f in faculties:
            selected = 'selected' if f['id'] in assigned_ids else ''
            faculty_options += f'<option value="{f["id"]}" {selected}>{display_text(f["faculty_name"], "")} ({display_text(f["department"], "")})</option>'

        html = self.render_template('edit_event',
            title='Edit Event',
            event_id=event_id,
            event_name=escape_html(event['event_name'] or ''),
            event_date=escape_html(event['date'] or ''),
            conducted_by=escape_html(event['conducted_by'] or ''),
            participants=escape_html(event['participants'] or ''),
            description=escape_html(event['description'] or ''),
            remarks=escape_html(event['remarks'] or ''),
            current_sections_json=json.dumps(current_sections_list),
            slot_options=slot_options,
            type_options=type_options,
            hall_options=hall_options,
            faculty_options=faculty_options)

        self.send_html(html)

    def handle_edit_event_post(self, event_id, body):
        try:
            payload = self._parse_event_payload(json.loads(body))
        except json.JSONDecodeError:
            self.send_json({'success': False, 'message': 'Invalid JSON'}, 400)
            return
        except ValueError as exc:
            self.send_json({'success': False, 'message': str(exc)}, 400)
            return

        db = get_db()
        try:
            self._validate_event_references(db, payload, event_id)
            conflicts = self._find_conflicts(db, payload, event_id)
            if conflicts:
                self.send_json({'success': False, 'message': 'Conflicts detected', 'conflicts': conflicts}, 409)
                return

            db.execute('''
                UPDATE event SET
                    event_name = ?, event_type = ?, description = ?, conducted_by = ?,
                    date = ?, timeslot_id = ?, hall_id = ?,
                    participants = ?, remarks = ?
                WHERE id = ?
            ''', (
                payload['event_name'],
                payload['event_type'],
                payload['description'],
                payload['conducted_by'],
                payload['date'],
                payload['timeslot_id'],
                payload['hall_id'],
                payload['participants'],
                payload['remarks'],
                event_id
            ))

            db.execute('DELETE FROM event_branch_section WHERE event_id = ?', (event_id,))
            for section in payload['sections']:
                db.execute('''
                    INSERT INTO event_branch_section (event_id, branch_id, section_id)
                    VALUES (?, ?, ?)
                ''', (event_id, section['branch_id'], section['section_id']))

            db.execute('DELETE FROM event_faculty WHERE event_id = ?', (event_id,))
            for faculty_id in payload['faculty_ids']:
                db.execute('INSERT INTO event_faculty (event_id, faculty_id) VALUES (?, ?)',
                          (event_id, faculty_id))

            db.commit()
        except ValueError as exc:
            db.rollback()
            self.send_json({'success': False, 'message': str(exc)}, 404 if str(exc) == 'Event not found' else 400)
            return
        except sqlite3.Error as exc:
            db.rollback()
            self.send_json({'success': False, 'message': f'Database error: {exc}'}, 500)
            return
        finally:
            db.close()

        self.send_json({'success': True})

    def handle_delete_event(self, event_id):
        db = get_db()
        if not db.execute('SELECT id FROM event WHERE id = ?', (event_id,)).fetchone():
            db.close()
            self.send_error(404)
            return
        db.execute('DELETE FROM event_branch_section WHERE event_id = ?', (event_id,))
        db.execute('DELETE FROM event_faculty WHERE event_id = ?', (event_id,))
        db.execute('DELETE FROM event WHERE id = ?', (event_id,))
        db.commit()
        db.close()
        self.send_redirect('/events', 'Event deleted successfully!')

    def handle_api_get(self, path, query):
        if path == '/api/branches':
            db = get_db()
            branches = db.execute('SELECT id, name FROM branch ORDER BY name').fetchall()
            result = [{'id': b['id'], 'name': b['name']} for b in branches]
            db.close()
            self.send_json(result)

        elif path.startswith('/api/sections/'):
            try:
                branch_id = int(path.split('/')[-1])
            except ValueError:
                self.send_json({'error': 'Invalid branch id'}, 400)
                return
            db = get_db()
            sections = db.execute('SELECT id, section_name FROM section WHERE branch_id = ? ORDER BY section_name',
                                  (branch_id,)).fetchall()
            result = [{'id': s['id'], 'name': s['section_name']} for s in sections]
            db.close()
            self.send_json(result)

        elif path == '/api/free-slots-multiple':
            date_val = query.get('date', [''])[0]
            branch_ids_str = query.get('branch_ids', [''])[0]
            section_ids_str = query.get('section_ids', [''])[0]

            if not all([date_val, branch_ids_str, section_ids_str]):
                self.send_json({'error': 'Missing parameters'}, 400)
                return

            try:
                branch_ids = [int(x) for x in branch_ids_str.split(',') if x]
                section_ids = [int(x) for x in section_ids_str.split(',') if x]
            except ValueError:
                self.send_json({'error': 'Invalid IDs'}, 400)
                return

            if not branch_ids or not section_ids:
                self.send_json({'error': 'No valid IDs provided'}, 400)
                return

            db = get_db()
            all_slots = db.execute('SELECT * FROM time_slot ORDER BY period_no').fetchall()

            placeholders = ','.join('?' for _ in section_ids)
            occupied = db.execute(f'''
                SELECT DISTINCT timeslot_id FROM event e
                JOIN event_branch_section ebs ON e.id = ebs.event_id
                WHERE e.date = ? AND ebs.section_id IN ({placeholders})
            ''', [date_val] + section_ids).fetchall()

            occupied_ids = [row['timeslot_id'] for row in occupied]

            free_slots = []
            for slot in all_slots:
                if slot['id'] not in occupied_ids:
                    free_slots.append({
                        'id': slot['id'],
                        'period_no': slot['period_no'],
                        'start_time': slot['start_time'],
                        'end_time': slot['end_time'],
                        'display': f"Period {slot['period_no']}: {slot['start_time']} - {slot['end_time']}"
                    })

            db.close()
            self.send_json(free_slots)

        elif path == '/api/available-faculty':
            date_val = query.get('date', [''])[0]
            timeslot_id = query.get('timeslot_id', [''])[0]

            if not all([date_val, timeslot_id]):
                self.send_json({'error': 'Missing parameters'}, 400)
                return

            db = get_db()
            event_day = datetime.strptime(date_val, '%Y-%m-%d').weekday()
            all_faculty = db.execute('SELECT id, faculty_name, department FROM faculty ORDER BY faculty_name').fetchall()

            busy_ids = set()
            unavailable = db.execute('''
                SELECT faculty_id FROM faculty_availability
                WHERE day_of_week = ? AND timeslot_id = ? AND is_available = 0
            ''', (event_day, timeslot_id)).fetchall()
            for u in unavailable:
                busy_ids.add(u['faculty_id'])

            events = db.execute('''
                SELECT ef.faculty_id FROM event_faculty ef
                JOIN event e ON ef.event_id = e.id
                WHERE e.date = ? AND e.timeslot_id = ?
            ''', (date_val, timeslot_id)).fetchall()

            for e in events:
                busy_ids.add(e['faculty_id'])

            available = []
            for f in all_faculty:
                if f['id'] not in busy_ids:
                    available.append({
                        'id': f['id'],
                        'name': f['faculty_name'],
                        'department': f['department']
                    })

            db.close()
            self.send_json(available)

        elif path == '/api/available-halls':
            date_val = query.get('date', [''])[0]
            timeslot_id = query.get('timeslot_id', [''])[0]

            if not all([date_val, timeslot_id]):
                self.send_json({'error': 'Missing parameters'}, 400)
                return

            db = get_db()
            all_halls = db.execute('SELECT id, hall_name, capacity FROM hall ORDER BY capacity DESC').fetchall()

            busy_ids = set()
            events = db.execute('''
                SELECT hall_id FROM event WHERE date = ? AND timeslot_id = ?
            ''', (date_val, timeslot_id)).fetchall()

            for e in events:
                busy_ids.add(e['hall_id'])

            available = []
            for h in all_halls:
                if h['id'] not in busy_ids:
                    available.append({
                        'id': h['id'],
                        'name': h['hall_name'],
                        'capacity': h['capacity']
                    })

            db.close()
            self.send_json(available)

        elif path.startswith('/api/events-by-date/'):
            date_str = path.split('/')[-1]
            db = get_db()
            events = db.execute('''
                SELECT e.id, e.event_name, e.event_type,
                       ts.period_no, ts.start_time, ts.end_time, h.hall_name,
                       GROUP_CONCAT(DISTINCT b.name || '-' || s.section_name) as branch_sections,
                       GROUP_CONCAT(DISTINCT f.faculty_name) as faculty_names
                FROM event e
                JOIN time_slot ts ON e.timeslot_id = ts.id
                JOIN hall h ON e.hall_id = h.id
                LEFT JOIN event_branch_section ebs ON e.id = ebs.event_id
                LEFT JOIN branch b ON ebs.branch_id = b.id
                LEFT JOIN section s ON ebs.section_id = s.id
                LEFT JOIN event_faculty ef ON e.id = ef.event_id
                LEFT JOIN faculty f ON ef.faculty_id = f.id
                WHERE e.date = ?
                GROUP BY e.id
                ORDER BY ts.period_no
            ''', (date_str,)).fetchall()
            db.close()

            result = []
            for e in events:
                result.append({
                    'id': e['id'],
                    'name': e['event_name'],
                    'type': e['event_type'],
                    'branch_sections': e['branch_sections'] or 'N/A',
                    'timeslot': f"Period {e['period_no']}: {e['start_time']} - {e['end_time']}",
                    'hall': e['hall_name'],
                    'faculty': e['faculty_names'] or 'Not Assigned'
                })

            self.send_json(result)

        else:
            self.send_error(404)

    def handle_check_conflicts(self, body):
        try:
            data = json.loads(body)
            payload = self._parse_event_payload(data)
        except json.JSONDecodeError:
            self.send_json({'conflicts': []}, 400)
            return
        except ValueError as exc:
            self.send_json({'conflicts': [{'type': 'validation', 'message': str(exc)}]}, 400)
            return

        event_id = data.get('event_id') or None
        if event_id:
            try:
                event_id = int(event_id)
            except (TypeError, ValueError):
                self.send_json({'conflicts': [{'type': 'validation', 'message': 'Event ID must be a valid number'}]}, 400)
                return

        db = get_db()
        try:
            self._validate_event_references(db, payload, event_id)
            conflicts = self._find_conflicts(db, payload, event_id)
        except ValueError as exc:
            self.send_json({'conflicts': [{'type': 'validation', 'message': str(exc)}]}, 400)
            return
        finally:
            db.close()
        self.send_json({'conflicts': conflicts})

    def handle_faculty_list(self, query=None):
        query = query or {}
        db = get_db()

        search_filter = (query.get('q', [''])[0] or '').strip()
        department_filter = (query.get('department', [''])[0] or '').strip()
        departments = ['CSE', 'CSE-AI', 'ECE', 'EEE', 'Mechanical', 'Civil', 'IT', 'General']
        if department_filter and department_filter not in departments:
            department_filter = ''

        clauses = []
        params = []
        if search_filter:
            clauses.append('(f.faculty_name LIKE ? OR f.department LIKE ? OR f.email LIKE ? OR f.phone LIKE ?)')
            like = f'%{search_filter}%'
            params.extend([like, like, like, like])
        if department_filter:
            clauses.append('f.department = ?')
            params.append(department_filter)

        where_clause = ''
        if clauses:
            where_clause = ' WHERE ' + ' AND '.join(clauses)

        faculties = db.execute(f'''
            SELECT f.*,
                   (SELECT COUNT(*) FROM event_faculty ef WHERE ef.faculty_id = f.id) as event_count
            FROM faculty f
            {where_clause}
            ORDER BY f.department, f.faculty_name
        ''', params).fetchall()
        db.close()

        cards = ''
        for f in faculties:
            search_text = escape_html(f"{f['faculty_name']} {f['department']} {f['email'] or ''} {f['phone'] or ''}").lower()
            cards += f'''<div class="col-lg-6 faculty-card-container" data-department="{escape_html(f['department'])}" data-search="{search_text}">
                <div class="card h-100">
                    <div class="card-body">
                        <div class="d-flex justify-content-between align-items-start gap-3">
                            <div>
                                <div class="d-flex align-items-center gap-2 mb-2">
                                    <h5 class="card-title mb-0">{display_text(f['faculty_name'], '')}</h5>
                                    <span class="badge bg-secondary">{display_text(f['department'], '')}</span>
                                </div>
                                <p class="text-muted mb-2">{display_text(f['email'], 'No email provided')}</p>
                                <p class="text-muted mb-0 small">{display_text(f['phone'], 'No phone number')}</p>
                            </div>
                            <div class="dropdown">
                                <button class="btn btn-sm btn-outline-secondary dropdown-toggle" data-bs-toggle="dropdown">Actions</button>
                                <ul class="dropdown-menu">
                                    <li><a class="dropdown-item" href="/faculty/{f['id']}/edit">Edit</a></li>
                                    <li><a class="dropdown-item" href="/faculty/{f['id']}/availability">Set Availability</a></li>
                                    <li>
                                        <form action="/faculty/{f['id']}/delete" method="POST" onsubmit="return confirm('Delete this faculty?')">
                                            <button type="submit" class="dropdown-item text-danger">Delete</button>
                                        </form>
                                    </li>
                                </ul>
                            </div>
                        </div>
                        <div class="mt-3 d-flex justify-content-between align-items-center">
                            <span class="result-pill">{f['event_count']} assigned event(s)</span>
                            <span class="text-muted small">Ready for scheduling</span>
                        </div>
                    </div>
                </div>
            </div>'''

        if not faculties:
            cards = '''<div class="col-12">
                <div class="card">
                    <div class="card-body text-center py-5">
                        <p class="text-muted mb-3">No faculty members matched the current filters.</p>
                        <a href="/faculty/add" class="btn btn-primary">Add Faculty</a>
                    </div>
                </div>
            </div>'''

        dept_options = '<option value="">All departments</option>'
        for dept in departments:
            selected = 'selected' if dept == department_filter else ''
            dept_options += f'<option value="{dept}" {selected}>{dept}</option>'

        html = self.render_template('faculty_list',
            title='Faculty',
            faculty_cards=cards,
            search_filter=escape_html(search_filter),
            department_filter=escape_html(department_filter),
            dept_options=dept_options,
            len_faculties=len(faculties))
        self.send_html(html)

    def handle_faculty_form(self, faculty_id=None):
        db = get_db()
        faculty = None
        if faculty_id:
            faculty = db.execute('SELECT * FROM faculty WHERE id = ?', (faculty_id,)).fetchone()
            if not faculty:
                db.close()
                self.send_error(404)
                return

        departments = ['CSE', 'CSE-AI', 'ECE', 'EEE', 'Mechanical', 'Civil', 'IT', 'General']
        dept_options = ''
        for dept in departments:
            selected = 'selected' if faculty and faculty['department'] == dept else ''
            dept_options += f'<option value="{dept}" {selected}>{dept}</option>'

        db.close()

        html = self.render_template('faculty_form',
            title='Edit Faculty' if faculty_id else 'Add Faculty',
            faculty_name=escape_html(faculty['faculty_name']) if faculty else '',
            email=escape_html(faculty['email']) if faculty else '',
            phone=escape_html(faculty['phone']) if faculty else '',
            dept_options=dept_options)

        self.send_html(html)

    def handle_faculty_post(self, body, faculty_id=None):
        data = parse_form(body)
        faculty_name = data.get('faculty_name', '').strip()
        department = data.get('department', '').strip()
        email = data.get('email', '').strip() or None
        phone = data.get('phone', '').strip()

        if not faculty_name or not department:
            self.send_html(self.render_template('faculty_form',
                title='Edit Faculty' if faculty_id else 'Add Faculty',
                faculty_name=escape_html(faculty_name),
                email=escape_html(email or ''),
                phone=escape_html(phone),
                dept_options=''.join(f'<option value="{dept}" {"selected" if dept == department else ""}>{dept}</option>'
                                     for dept in ['CSE', 'CSE-AI', 'ECE', 'EEE', 'Mechanical', 'Civil', 'IT', 'General'])),
                status=400)
            return

        db = get_db()
        try:
            if faculty_id:
                cursor = db.execute('UPDATE faculty SET faculty_name = ?, department = ?, email = ?, phone = ? WHERE id = ?',
                           (faculty_name, department, email, phone, faculty_id))
                if cursor.rowcount == 0:
                    self.send_error(404)
                    return
            else:
                db.execute('INSERT INTO faculty (faculty_name, department, email, phone) VALUES (?, ?, ?, ?)',
                           (faculty_name, department, email, phone))
            db.commit()
        except sqlite3.IntegrityError:
            db.rollback()
            self.send_error(400, 'Faculty email already exists.')
            return
        finally:
            db.close()
        self.send_redirect('/faculty', 'Faculty saved successfully!')

    def handle_delete_faculty(self, faculty_id):
        db = get_db()
        if not db.execute('SELECT id FROM faculty WHERE id = ?', (faculty_id,)).fetchone():
            db.close()
            self.send_error(404)
            return
        assigned_count = db.execute('SELECT COUNT(*) FROM event_faculty WHERE faculty_id = ?', (faculty_id,)).fetchone()[0]
        if assigned_count:
            db.close()
            self.send_redirect('/faculty', 'Cannot delete faculty assigned to events.')
            return
        db.execute('DELETE FROM faculty_availability WHERE faculty_id = ?', (faculty_id,))
        db.execute('DELETE FROM faculty_busy_slot WHERE faculty_id = ?', (faculty_id,))
        db.execute('DELETE FROM faculty WHERE id = ?', (faculty_id,))
        db.commit()
        db.close()
        self.send_redirect('/faculty', 'Faculty deleted!')

    def handle_faculty_availability(self, faculty_id):
        db = get_db()
        faculty = db.execute('SELECT * FROM faculty WHERE id = ?', (faculty_id,)).fetchone()
        if not faculty:
            db.close()
            self.send_error(404)
            return

        time_slots = db.execute('SELECT * FROM time_slot ORDER BY period_no').fetchall()
        
        availability = db.execute('''
            SELECT day_of_week, timeslot_id 
            FROM faculty_availability 
            WHERE faculty_id = ? AND is_available = 0
        ''', (faculty_id,)).fetchall()
        
        unavailable_slots = set()
        for a in availability:
            key = f"{a['day_of_week']}_{a['timeslot_id']}"
            unavailable_slots.add(key)

        availability_rows = ''
        
        for slot in time_slots:
            availability_rows += f'<div class="availability-row" data-period="{slot["period_no"]}">'
            availability_rows += f'<div class="availability-label">P{slot["period_no"]}<br><small>{slot["start_time"]}-{slot["end_time"]}</small></div>'
            
            for day in range(7):
                key = f"{day}_{slot['id']}"
                checked = 'checked' if key in unavailable_slots else ''
                availability_rows += f'''
                    <div class="availability-cell">
                        <input type="checkbox" name="availability_{day}_{slot['id']}" {checked}>
                    </div>
                '''
            availability_rows += '</div>'

        db.close()

        html = self.render_template('faculty_availability',
            title='Set Faculty Availability',
            faculty_name=escape_html(faculty['faculty_name']),
            availability_rows=availability_rows)

        self.send_html(html)

    def handle_faculty_availability_post(self, faculty_id, body):
        db = get_db()
        
        if not db.execute('SELECT id FROM faculty WHERE id = ?', (faculty_id,)).fetchone():
            db.close()
            self.send_error(404)
            return

        try:
            data = parse_form(body)
            db.execute('DELETE FROM faculty_availability WHERE faculty_id = ?', (faculty_id,))
            
            unavailable_slots = set()
            for key, value in data.items():
                if key.startswith('availability_'):
                    parts = key.split('_')
                    if len(parts) == 3:
                        day = int(parts[1])
                        timeslot_id = int(parts[2])
                        if value == 'on':
                            unavailable_slots.add((day, timeslot_id))
            
            for day, timeslot_id in unavailable_slots:
                db.execute('''
                    INSERT INTO faculty_availability (faculty_id, day_of_week, timeslot_id, is_available)
                    VALUES (?, ?, ?, 0)
                ''', (faculty_id, day, timeslot_id))
            
            db.commit()
            
        except sqlite3.Error as exc:
            db.rollback()
            self.send_error(500, f'Database error: {exc}')
            return
        except Exception as exc:
            db.rollback()
            self.send_error(500, f'Error: {exc}')
            return
        finally:
            db.close()
            
        self.send_redirect('/faculty', 'Faculty availability updated successfully!')

    def handle_hall_list(self):
        db = get_db()
        halls = db.execute('''
            SELECT h.*,
                   (SELECT COUNT(*) FROM event WHERE hall_id = h.id) as event_count
            FROM hall h
            ORDER BY h.hall_name
        ''').fetchall()
        db.close()

        hall_rows = ''
        for h in halls:
            hall_rows += f'''<tr>
                <td>{display_text(h['hall_name'], '')}</td>
                <td>{display_text(h['capacity'], '')}</td>
                <td>{h['event_count']}</td>
                <td>
                    <a href="/halls/{h['id']}/edit" class="btn btn-sm btn-outline-primary">Edit</a>
                    <form action="/halls/{h['id']}/delete" method="POST" style="display: inline;" onsubmit="return confirm('Delete this hall?')">
                        <button type="submit" class="btn btn-sm btn-outline-danger">Delete</button>
                    </form>
                </td>
            </tr>'''

        html = self.render_template('hall_list', title='Halls', hall_rows=hall_rows)
        self.send_html(html)

    def handle_hall_form(self, hall_id=None):
        db = get_db()
        hall = None
        if hall_id:
            hall = db.execute('SELECT * FROM hall WHERE id = ?', (hall_id,)).fetchone()
            if not hall:
                db.close()
                self.send_error(404)
                return
        db.close()

        html = self.render_template('hall_form',
            title='Edit Hall' if hall_id else 'Add Hall',
            hall_name=escape_html(hall['hall_name']) if hall else '',
            capacity=escape_html(hall['capacity']) if hall else '')

        self.send_html(html)

    def handle_hall_post(self, body, hall_id=None):
        data = parse_form(body)
        hall_name = data.get('hall_name', '').strip()
        try:
            capacity = int(data.get('capacity', ''))
        except ValueError:
            self.send_error(400, 'Capacity must be a valid number.')
            return

        if not hall_name or capacity < 1:
            self.send_error(400, 'Hall name is required and capacity must be at least 1.')
            return

        db = get_db()
        try:
            if hall_id:
                cursor = db.execute('UPDATE hall SET hall_name = ?, capacity = ? WHERE id = ?',
                           (hall_name, capacity, hall_id))
                if cursor.rowcount == 0:
                    self.send_error(404)
                    return
            else:
                db.execute('INSERT INTO hall (hall_name, capacity) VALUES (?, ?)',
                           (hall_name, capacity))
            db.commit()
        except sqlite3.IntegrityError:
            db.rollback()
            self.send_error(400, 'Hall name already exists.')
            return
        finally:
            db.close()
        self.send_redirect('/halls', 'Hall saved successfully!')

    def handle_delete_hall(self, hall_id):
        db = get_db()
        if not db.execute('SELECT id FROM hall WHERE id = ?', (hall_id,)).fetchone():
            db.close()
            self.send_error(404)
            return
        assigned_count = db.execute('SELECT COUNT(*) FROM event WHERE hall_id = ?', (hall_id,)).fetchone()[0]
        if assigned_count:
            db.close()
            self.send_redirect('/halls', 'Cannot delete hall assigned to events.')
            return
        db.execute('DELETE FROM hall WHERE id = ?', (hall_id,))
        db.commit()
        db.close()
        self.send_redirect('/halls', 'Hall deleted!')

    def handle_send_reminders(self):
        success = send_tomorrow_event_reminders()
        if success:
            self.send_json({'success': True, 'message': 'Reminder emails sent successfully'})
        else:
            self.send_json({'success': False, 'message': 'Reminder emails could not be sent'}, 500)


def main():
    init_db()
    print("=" * 60)
    print("Event Scheduler started!")
    print(f"Server running on http://localhost:{PORT}")
    print("-" * 60)
    print("Default Login Credentials:")
    for username, password in VALID_CREDENTIALS.items():
        print(f"  Username: {username}, Password: {password}")
    print("=" * 60)
    print(f"Reminder emails enabled at {REMINDER_HOUR:02d}:{REMINDER_MINUTE:02d} daily")

    scheduler_thread = threading.Thread(target=run_daily_email_scheduler, name='event-reminder-scheduler', daemon=True)
    scheduler_thread.start()

    with ThreadedTCPServer(("", PORT), EventSchedulerHandler) as httpd:
        try:
            httpd.serve_forever()
        except KeyboardInterrupt:
            print("\nServer stopped.")


if __name__ == '__main__':
    main()